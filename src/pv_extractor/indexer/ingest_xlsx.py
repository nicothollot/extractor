"""Bulk loader (Loader 1) for a PV index export workbook.

Streams the export with openpyxl read_only mode (millions of rows). Only
file_path, size_bytes and modified_time are trusted; every other column is
re-derived from file_path via derive_record (CLAUDE.md rule 4 — the export's
own derived columns contain corruption such as '#NAME?' parent folders).
"""

from __future__ import annotations

import logging
import sqlite3
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

from pv_extractor.config import Config
from pv_extractor.indexer.db import add_scan_error, insert_records
from pv_extractor.indexer.derive import derive_record
from pv_extractor.logging_setup import log_event
from pv_extractor.models import FileRecord, ScanError

logger = logging.getLogger(__name__)

_TRUSTED_COLUMNS = ("file_path", "size_bytes", "modified_time")


def _parse_modified_time(value: object) -> datetime | None:
    if value is None or isinstance(value, datetime):
        return value
    return datetime.fromisoformat(str(value))


def ingest_xlsx(
    conn: sqlite3.Connection,
    xlsx_path: Path,
    config: Config,
    limit: int | None = None,
) -> int:
    """Load up to `limit` export rows into `files`; returns rows ingested.

    Rows whose file_path falls outside config.pv_root are recorded in
    scan_errors and skipped rather than aborting a multi-million-row load.
    """
    workbook = load_workbook(xlsx_path, read_only=True, data_only=True)
    try:
        rows = workbook.worksheets[0].iter_rows(values_only=True)
        header = next(rows, None)
        if header is None:
            raise ValueError(f"export workbook has no header row: {xlsx_path}")
        column = {str(name): i for i, name in enumerate(header) if name is not None}
        missing = [name for name in _TRUSTED_COLUMNS if name not in column]
        if missing:
            raise ValueError(f"export workbook is missing columns {missing}: {xlsx_path}")

        total = 0
        batch: list[FileRecord] = []
        for row in rows:
            if limit is not None and total + len(batch) >= limit:
                break
            file_path = row[column["file_path"]] if len(row) > column["file_path"] else None
            if not file_path:
                continue
            size = row[column["size_bytes"]]
            try:
                record = derive_record(
                    str(file_path),
                    size_bytes=int(size) if size is not None else None,
                    modified_time=_parse_modified_time(row[column["modified_time"]]),
                    config=config,
                )
            except ValueError as exc:
                add_scan_error(
                    conn,
                    ScanError(
                        path=str(file_path),
                        error_type=type(exc).__name__,
                        message=str(exc),
                        seen_at=datetime.now(),
                    ),
                )
                continue
            batch.append(record)
            if len(batch) >= config.indexer.batch_size:
                total += insert_records(conn, batch, config.indexer.batch_size)
                batch.clear()
                log_event(logger, "ingest_xlsx batch committed", xlsx=str(xlsx_path), rows_ingested=total)
        if batch:
            total += insert_records(conn, batch, config.indexer.batch_size)
            log_event(logger, "ingest_xlsx batch committed", xlsx=str(xlsx_path), rows_ingested=total)
        log_event(logger, "ingest_xlsx complete", xlsx=str(xlsx_path), rows_ingested=total)
        return total
    finally:
        workbook.close()
