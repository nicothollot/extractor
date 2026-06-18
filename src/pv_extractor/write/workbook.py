"""Workbook writer (D6).

NEVER touches the reference template in place: copy_template byte-copies it
(open_read -> guarded_open_write) into the run directory and every write
happens on the copy. Before any row is appended the Index sheet's row-2
headers are asserted against the compiled schema (HeaderDriftError aborts
the run — writing by column index into a drifted template would corrupt it).

Cells are written by column INDEX from the schema JSON, never resolved by
header string at write time. Values are plain (no formulas ever — a string
that looks like a formula is forced to stay a string). Review Flags rows
dedupe on (memo_id, flag_description): the reference file carries duplicate
rows; this writer never adds more.
"""

from __future__ import annotations

import logging
from datetime import date
from pathlib import Path

import openpyxl

from pv_extractor.io_guard import assert_write_allowed, guarded_open_write, open_read
from pv_extractor.logging_setup import log_event
from pv_extractor.models import FieldHit, QaStatus, ReviewFlag, SchemaField

logger = logging.getLogger(__name__)

INDEX_SHEET = "Index"
FLAGS_SHEET = "Review Flags"
RUNLOG_SHEET = "Run Log"

_HEADER_ROW = 2
_FIRST_DATA_ROW = 4

RUN_LOG_COLUMNS = (
    "Run ID", "Run Date", "Memos Processed", "Assets Extracted", "QA Pass",
    "QA Pass with Flags", "QA Fail", "Records Added to Index", "Total Flags",
    "Reviewer Attention Items", "Run Duration (mins)", "Batch Sessions", "Notes",
)

FLAG_COLUMNS = (
    "Run ID", "Memo ID", "Source Filename", "Fund Manager", "Portfolio Company",
    "Valuation Date", "QA Status", "Flag #", "Flag Description", "Flag Category",
    "Reviewer Attention (Y/N)", "Resolved (Y/N)", "Resolution Notes",
)


class HeaderDriftError(RuntimeError):
    """Template headers no longer match the compiled schema — hard abort."""


def copy_template(template_path: str | Path, dest_path: str | Path, pv_root: str) -> Path:
    """Byte-copy the template to the run directory through the io guard."""
    dest = Path(dest_path)
    with open_read(template_path) as src, guarded_open_write(dest, pv_root, mode="wb") as out:
        while chunk := src.read(1 << 20):
            out.write(chunk)
    log_event(logger, "template copied", template=str(template_path), dest=str(dest))
    return dest


def _cell_value(value: object) -> object:
    """Workbook representation: booleans 'Yes'/'No' (reference convention),
    dates ISO strings, formula-looking strings defused, numbers as numbers."""
    if isinstance(value, bool):
        return "Yes" if value else "No"
    if isinstance(value, date):
        return value.isoformat()
    return value


class WorkbookWriter:
    """All writes for one run, against one template COPY."""

    def __init__(self, path: str | Path, schema_fields: list[SchemaField], pv_root: str) -> None:
        self.path = Path(path)
        self.pv_root = pv_root
        self.schema_fields = schema_fields
        self.workbook = openpyxl.load_workbook(self.path)
        self.assert_headers()
        self._existing_flag_keys = self._load_flag_keys()

    # ------------------------------------------------------------------
    # header drift gate
    # ------------------------------------------------------------------

    def assert_headers(self) -> None:
        sheet = self.workbook[INDEX_SHEET]
        for field in self.schema_fields:
            actual = sheet.cell(row=_HEADER_ROW, column=field.col_index).value
            if (actual or "") != field.header:
                raise HeaderDriftError(
                    f"Index header drift at column {field.col_index}: template has "
                    f"{actual!r}, schema expects {field.header!r} — aborting before any write"
                )
        flags_sheet = self.workbook[FLAGS_SHEET]
        actual_flags = tuple(
            flags_sheet.cell(row=1, column=i + 1).value or "" for i in range(len(FLAG_COLUMNS))
        )
        if actual_flags != FLAG_COLUMNS:
            raise HeaderDriftError(
                f"Review Flags header drift: {actual_flags!r} != {FLAG_COLUMNS!r}"
            )

    # ------------------------------------------------------------------
    # Index sheet
    # ------------------------------------------------------------------

    def _next_index_row(self) -> int:
        sheet = self.workbook[INDEX_SHEET]
        for row in range(sheet.max_row, _FIRST_DATA_ROW - 1, -1):
            if any(
                sheet.cell(row=row, column=col).value is not None
                for col in (1, 2, 3)
            ):
                return row + 1
        return _FIRST_DATA_ROW

    def append_index_row(self, hits: list[FieldHit]) -> int:
        """One memo-asset row, written by schema column index. Returns the
        worksheet row number used."""
        sheet = self.workbook[INDEX_SHEET]
        row = self._next_index_row()
        for hit in hits:
            if hit.value is None:
                continue
            cell = sheet.cell(row=row, column=hit.col_index)
            value = _cell_value(hit.value)
            cell.value = value
            if isinstance(value, str) and value.startswith("="):
                cell.data_type = "s"  # no formulas ever
        return row

    def find_index_row(self, row_memo_id: str) -> int | None:
        """Worksheet row of an existing memo-asset row (Memo ID = column 1)."""
        sheet = self.workbook[INDEX_SHEET]
        for row in range(_FIRST_DATA_ROW, sheet.max_row + 1):
            value = sheet.cell(row=row, column=1).value
            if value is not None and str(value) == row_memo_id:
                return row
        return None

    def update_cell(self, row_memo_id: str, col_index: int, value: object) -> int:
        """Phase-4 review-queue edit seam: overwrite ONE cell of an existing
        row, addressed by Memo ID + schema column index (never by header
        lookup). Same value conventions as append_index_row ('Yes'/'No'
        booleans, ISO dates, formula-looking strings defused). Returns the
        worksheet row; raises KeyError when the memo row does not exist."""
        row = self.find_index_row(row_memo_id)
        if row is None:
            raise KeyError(f"no Index row for memo id {row_memo_id!r}")
        sheet = self.workbook[INDEX_SHEET]
        cell = sheet.cell(row=row, column=col_index)
        rendered = _cell_value(value)
        cell.value = rendered
        if isinstance(rendered, str) and rendered.startswith("="):
            cell.data_type = "s"  # no formulas ever
        log_event(
            logger, "index cell updated", memo_id=row_memo_id,
            col_index=col_index, row=row,
        )
        return row

    def resolve_flag(
        self, memo_id: str, description: str, *, resolved: bool, note: str | None = None
    ) -> bool:
        """Phase-4 review-queue seam: mark one Review Flags row Resolved
        (Y/N) and set its Resolution Notes. Matching is on the same
        (memo_id, description) key the dedupe uses. Returns True when a
        matching row was found."""
        sheet = self.workbook[FLAGS_SHEET]
        resolved_col = FLAG_COLUMNS.index("Resolved (Y/N)") + 1
        notes_col = FLAG_COLUMNS.index("Resolution Notes") + 1
        for row in range(2, sheet.max_row + 1):
            row_memo = sheet.cell(row=row, column=2).value
            row_desc = sheet.cell(row=row, column=9).value
            if row_memo is None or row_desc is None:
                continue
            if str(row_memo) == memo_id and str(row_desc) == description:
                sheet.cell(row=row, column=resolved_col).value = "Y" if resolved else "N"
                if note is not None:
                    sheet.cell(row=row, column=notes_col).value = note
                log_event(
                    logger, "review flag resolved", memo_id=memo_id,
                    resolved=resolved, row=row,
                )
                return True
        return False

    def find_prior_row(
        self, portfolio_company: str | None, fund_name: str | None, before: date
    ) -> dict[str, object] | None:
        """Latest existing Index row for the same asset (Portfolio Company +
        Fund Name match) with Valuation Date before `before` — the QoQ
        continuity baseline (D5)."""
        if not portfolio_company:
            return None
        sheet = self.workbook[INDEX_SHEET]
        col_by_header = {field.header: field.col_index for field in self.schema_fields}
        company_col = col_by_header["Portfolio Company"]
        fund_col = col_by_header["Fund Name"]
        date_col = col_by_header["Valuation Date"]

        best: tuple[date, int] | None = None
        for row in range(_FIRST_DATA_ROW, sheet.max_row + 1):
            company = sheet.cell(row=row, column=company_col).value
            if not company or str(company).strip().lower() != portfolio_company.strip().lower():
                continue
            fund = sheet.cell(row=row, column=fund_col).value
            if fund_name and fund and str(fund).strip().lower() != fund_name.strip().lower():
                continue
            raw_date = sheet.cell(row=row, column=date_col).value
            try:
                row_date = (
                    raw_date if isinstance(raw_date, date) else date.fromisoformat(str(raw_date))
                )
            except (TypeError, ValueError):
                continue
            if row_date < before and (best is None or row_date > best[0]):
                best = (row_date, row)
        if best is None:
            return None
        row = best[1]
        return {
            field.header: sheet.cell(row=row, column=field.col_index).value
            for field in self.schema_fields
        }

    # ------------------------------------------------------------------
    # Review Flags sheet
    # ------------------------------------------------------------------

    def _load_flag_keys(self) -> set[tuple[str, str]]:
        sheet = self.workbook[FLAGS_SHEET]
        keys: set[tuple[str, str]] = set()
        for row in sheet.iter_rows(min_row=2, values_only=True):
            memo_id, description = row[1], row[8]
            if memo_id and description:
                keys.add((str(memo_id), str(description)))
        return keys

    def append_review_flags(
        self,
        *,
        run_id: str,
        memo_id: str,
        source_filename: str,
        fund_manager: str | None,
        portfolio_company: str | None,
        valuation_date: date | None,
        qa_status: QaStatus,
        flags: list[ReviewFlag],
    ) -> int:
        """Append deduped flag rows; returns how many were actually added."""
        sheet = self.workbook[FLAGS_SHEET]
        added = 0
        for number, flag in enumerate(flags, start=1):
            key = (memo_id, flag.description)
            if key in self._existing_flag_keys:
                continue
            self._existing_flag_keys.add(key)
            sheet.append(
                (
                    run_id,
                    memo_id,
                    source_filename,
                    fund_manager,
                    portfolio_company,
                    valuation_date.isoformat() if valuation_date else None,
                    qa_status.value,
                    number,
                    flag.description,
                    flag.category,
                    "Y" if flag.reviewer_attention else "N",
                    "N",
                    None,
                )
            )
            added += 1
        return added

    # ------------------------------------------------------------------
    # Run Log sheet
    # ------------------------------------------------------------------

    def append_run_log(self, values: dict[str, object]) -> None:
        """One row per run; `values` keys must be RUN_LOG_COLUMNS members
        (Batch Sessions stays empty until Phase 3)."""
        unknown = set(values) - set(RUN_LOG_COLUMNS)
        if unknown:
            raise ValueError(f"unknown Run Log columns: {sorted(unknown)}")
        sheet = self.workbook[RUNLOG_SHEET]
        sheet.append(tuple(_cell_value(values.get(column)) for column in RUN_LOG_COLUMNS))

    def save(self) -> None:
        assert_write_allowed(self.path, self.pv_root)
        self.workbook.save(self.path)
        log_event(logger, "workbook saved", path=str(self.path))
