"""Compile the master index workbook (header rows 1-3) into schema artifacts.

Reads ``reference/master_index_v4.xlsx`` sheet "Index" — row 1 band names
(carried forward across each band), row 2 verbatim field headers, row 3
descriptions (the authoritative spec) — and emits two byte-stable JSON
artifacts under ``schema/``:

* ``master_schema.json``  — one :class:`SchemaField` per column, in column order.
* ``band_routing.json``   — methodology -> band-name routing derived from the
  "Primary Methodology" controlled vocab plus band/field names (never a
  hardcoded methodology list).

All dtype/vocab inference is a deterministic, pure function of workbook
content. Refinements made to the baseline inference rules after inspecting
the workbook (each deterministic and documented):

1. Slash-list enums ("A / B / C") require >= 3 items, each 1-4 words starting
   with an uppercase letter or digit. This rejects the two-item phrase
   "NOI / Cap Rate" (Cap Implied Asset Value description) and prose lists with
   lowercase items like "Change driven by EBITDA / NOI / revenue performance".
2. Explicit "vocab:" captures strip one trailing "." — a list at the end of a
   cell terminates with "." rather than ". " (e.g. Dev Stage).
3. A first sentence that is a comma list of short capitalized labels but
   contains "etc" ("LTM, NTM, FY+1, FY+2 etc.") marks a free-label STRING
   field and stops numeric header inference — otherwise "Mult Basis Year"
   would fall through to the ends-with-"Year" -> integer rule.
4. "same vocab" references resolve against ALL fields with an explicit vocab,
   not only already-compiled (preceding) ones: first by normalized-header
   substring of the normalized description, then by header-token subset
   (covers "Entry Multiple Metric" -> "Mult Metric", which sits LATER in the
   sheet and whose header is not a contiguous substring of the description),
   finally by the nearest preceding field that has a vocab.
5. The "Rate" numeric cue matches on a word boundary so "Yield Calibration
   Risk Rating" stays a string while "FX Rate (Current)" is a number.
6. Headers containing "(Mult" — "WF Attach/Detach Point (Mult of EBITDA)" —
   are multiple_x; the bare "(x)"/"Multiple" cues miss them.
"""

from __future__ import annotations

import json
import logging
import re
from pathlib import Path

from openpyxl import load_workbook

from pv_extractor.io_guard import guarded_open_write
from pv_extractor.logging_setup import log_event
from pv_extractor.models import SchemaField
from pv_extractor.normalize import normalize_text

logger = logging.getLogger(__name__)

SHEET_NAME = "Index"
MASTER_SCHEMA_FILENAME = "master_schema.json"
BAND_ROUTING_FILENAME = "band_routing.json"

_SLOT_RE = re.compile(r"^(TC|TX|CS)(\d{2})\s+(.*)$")
_SAME_VOCAB_RE = re.compile(r"same (?:controlled )?vocab", re.IGNORECASE)
_EXPLICIT_VOCAB_RE = re.compile(r"vocab\s*:\s*(.+)", re.IGNORECASE | re.DOTALL)
_NEW_PREFIX_RE = re.compile(r"^NEW\b\.?\s*")
_COMPACT_NUMERIC_RE = re.compile(r"^(.*?)\s?(\d(?:/\d)+)$")
_ETC_RE = re.compile(r"\betc\b", re.IGNORECASE)
_YN_RE = re.compile(r"\bY/N\b")
_DATE_HEADER_RE = re.compile(r"\bDate$")
_RATE_WORD_RE = re.compile(r"\bRate\b")

_ENUM_ITEM_MAX_WORDS = 4
_MIN_VOCAB_ITEMS = 2
_MIN_SLASH_ITEMS = 3
_COMMA_LIST_MAX_CHARS = 90
_MIN_REF_HEADER_TOKENS = 2
_REQUIRED_BAND = "IDENTIFICATION"
_FLAG_BAND_MARKER = "FLAG"
_NUMERIC_LEFTOVER_HEADERS = frozenset({"DSCR", "DPI"})
_PRIMARY_METHODOLOGY_HEADER = "Primary Methodology"
_METHODOLOGY_BAND_PREFIX = "METHODOLOGY:"
_TRADING_COMPS_MARKER = "TRADING COMPS"
_TRANSACTION_COMPS_MARKER = "TRANSACTION COMPS"
_ROUTING_DERIVED_FROM = (
    "Primary Methodology / Secondary Methodology controlled vocab (METHODOLOGY ROUTING band)"
)


# --------------------------------------------------------------------------
# Workbook reading
# --------------------------------------------------------------------------


def _read_header_rows(workbook_path: Path) -> tuple[list[str], list[str], list[str]]:
    """Return (bands, headers, descriptions), band names carried forward."""
    wb = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        ws = wb[SHEET_NAME]
        row1, row2, row3 = list(
            ws.iter_rows(min_row=1, max_row=3, min_col=1, max_col=ws.max_column, values_only=True)
        )
    finally:
        wb.close()

    # Trim trailing columns with no field header, then require contiguity.
    raw = list(zip(row1, row2, row3, strict=True))
    while raw and (raw[-1][1] is None or not str(raw[-1][1]).strip()):
        raw.pop()
    bands: list[str] = []
    headers: list[str] = []
    descs: list[str] = []
    band = ""
    for col, (b, h, d) in enumerate(raw, start=1):
        if b is not None and str(b).strip():
            band = str(b).strip()
        if h is None or not str(h).strip():
            raise ValueError(f"column {col}: empty field header inside the Index sheet")
        bands.append(band)
        headers.append(str(h))
        descs.append("" if d is None else str(d))
    return bands, headers, descs


# --------------------------------------------------------------------------
# Controlled-vocab inference
# --------------------------------------------------------------------------


def _first_sentence(desc: str) -> str:
    """Description minus a leading 'NEW.'/'NEW', cut at the first '. ',
    with a trailing '.' stripped."""
    text = _NEW_PREFIX_RE.sub("", desc.strip())
    cut = text.find(". ")
    if cut != -1:
        text = text[:cut]
    return text.rstrip(".").strip()


def _enum_item_ok(item: str) -> bool:
    words = item.split()
    return bool(item) and len(words) <= _ENUM_ITEM_MAX_WORDS and (item[0].isupper() or item[0].isdigit())


def _explicit_vocab(desc: str) -> list[str] | None:
    """'... vocab: A, B, C.' -> [A, B, C]; needs >= 2 items."""
    m = _EXPLICIT_VOCAB_RE.search(desc)
    if not m:
        return None
    captured = m.group(1)
    cut = captured.find(". ")
    if cut != -1:
        captured = captured[:cut]
    captured = captured.strip().rstrip(".")
    items = [item.strip() for item in captured.split(",") if item.strip()]
    return items if len(items) >= _MIN_VOCAB_ITEMS else None


def _slash_vocab(sentence: str) -> list[str] | None:
    """'A / B / C' (>= 3 short capitalized items); also tried on the text
    after a final ': ' ('Routing confidence: High / Medium / Low')."""
    candidates = []
    if ": " in sentence:
        candidates.append(sentence.rsplit(": ", 1)[1])
    candidates.append(sentence)
    for cand in candidates:
        items = [item.strip() for item in cand.split(" / ")]
        if len(items) >= _MIN_SLASH_ITEMS and all(_enum_item_ok(item) for item in items):
            return items
    return None


def _compact_vocab(sentence: str) -> list[str] | None:
    """'Level 1/2/3' -> ['Level 1', 'Level 2', 'Level 3']."""
    m = _COMPACT_NUMERIC_RE.match(sentence)
    if not m:
        return None
    prefix = m.group(1).strip()
    return [f"{prefix} {n}".strip() for n in m.group(2).split("/")]


def _comma_vocab(sentence: str) -> tuple[list[str] | None, bool]:
    """Comma list of >= 2 short capitalized items -> (vocab, False).
    A would-be list containing 'etc' -> (None, True): free-label string."""
    if "," not in sentence or sentence.lower().startswith("e.g."):
        return None, False
    items = [item.strip() for item in sentence.split(",")]
    if len(items) < _MIN_VOCAB_ITEMS or not all(_enum_item_ok(item) for item in items):
        return None, False
    if _ETC_RE.search(sentence):
        return None, True  # label list like "LTM, NTM, FY+1, FY+2 etc."
    if len(sentence) > _COMMA_LIST_MAX_CHARS:
        return None, False
    return items, False


def _vocab_state(desc: str) -> tuple[str | None, list[str] | None]:
    """Classify a description: ('ref', None) for 'same vocab' references,
    ('explicit', items) for resolved vocabs, ('label', None) for free-label
    comma lists, (None, None) otherwise."""
    if not desc:
        return None, None
    if _SAME_VOCAB_RE.search(desc):
        return "ref", None
    items = _explicit_vocab(desc)
    if items:
        return "explicit", items
    sentence = _first_sentence(desc)
    if not sentence:
        return None, None
    items = _slash_vocab(sentence) or _compact_vocab(sentence)
    if items:
        return "explicit", items
    items, is_label = _comma_vocab(sentence)
    if items:
        return "explicit", items
    if is_label:
        return "label", None
    return None, None


def _resolve_vocab_refs(
    states: list[tuple[str | None, list[str] | None]],
    headers: list[str],
    descs: list[str],
) -> list[list[str] | None]:
    """Second pass: copy vocabs into 'same vocab' reference fields."""
    resolved: list[list[str] | None] = [vocab for _, vocab in states]
    explicit = [i for i, (state, _) in enumerate(states) if state == "explicit"]

    def _pick(candidates: list[int]) -> int | None:
        if not candidates:
            return None
        # Longest (most specific) header wins; ties -> lowest column.
        return max(candidates, key=lambda j: (len(normalize_text(headers[j])), -j))

    for i, (state, _) in enumerate(states):
        if state != "ref":
            continue
        d_norm = normalize_text(descs[i])
        d_tokens = set(d_norm.split())
        source = _pick(
            [j for j in explicit if normalize_text(headers[j]) in d_norm]
        )
        if source is None:
            source = _pick(
                [
                    j
                    for j in explicit
                    if len(toks := set(normalize_text(headers[j]).split())) >= _MIN_REF_HEADER_TOKENS
                    and toks <= d_tokens
                ]
            )
        if source is None:  # nearest preceding field with a vocab
            source = next((j for j in range(i - 1, -1, -1) if resolved[j] is not None), None)
        if source is not None:
            resolved[i] = list(resolved[source] or [])
    return resolved


# --------------------------------------------------------------------------
# dtype inference
# --------------------------------------------------------------------------


def _numeric_dtype(header: str) -> tuple[str, str | None] | None:
    """Numeric dtype/unit cues on the (slot-stripped) header; first match wins."""
    if "(bps)" in header:
        return "basis_points", "bps"
    if "($M, local)" in header:
        return "number", "millions_local"
    if "($M" in header:  # ($M), ($M, USD)
        return "number", "USD_millions"
    if (
        "(x)" in header
        or "MOIC" in header
        or "EV/" in header
        or "/EBITDA" in header
        or "(Mult" in header
        or ("Multiple" in header and "Metric" not in header and "Rationale" not in header)
    ):
        return "multiple_x", "x"
    if "(yrs)" in header:
        return "years", "years"
    if "%" in header:
        return "percent", "percent"
    if header.endswith("Year") or "Vintage" in header:
        return "integer", None
    if header.startswith("Months"):
        return "number", "months"
    if header.endswith("Count"):
        return "integer", None
    if (
        header.endswith("Value")
        or _RATE_WORD_RE.search(header)
        or "Beta" in header
        or header in _NUMERIC_LEFTOVER_HEADERS
    ):
        return "number", None
    return None


def _build_field(
    col_index: int,
    band: str,
    header: str,
    desc: str,
    vocab: list[str] | None,
    is_label: bool,
) -> SchemaField:
    slot_group: str | None = None
    slot_number: int | None = None
    h = header
    m = _SLOT_RE.match(header)
    if m:
        slot_group, slot_number, h = m.group(1), int(m.group(2)), m.group(3)

    unit: str | None = None
    if vocab is not None:
        dtype = "enum"
    elif is_label:
        dtype = "string"
    elif _YN_RE.search(desc) or _FLAG_BAND_MARKER in band.upper() or h.endswith("Y/N"):
        dtype = "boolean"
    elif _DATE_HEADER_RE.search(h):
        dtype = "date"
    elif numeric := _numeric_dtype(h):
        dtype, unit = numeric
    elif desc.startswith("Numeric value"):
        dtype = "number"
    else:
        dtype = "string"

    return SchemaField(
        col_index=col_index,
        band=band,
        header=header,
        description=desc,
        dtype=dtype,
        controlled_vocab=vocab,
        unit=unit,
        slot_group=slot_group,
        slot_number=slot_number,
        required=band == _REQUIRED_BAND,
    )


# --------------------------------------------------------------------------
# Methodology -> band routing
# --------------------------------------------------------------------------


def _build_routing(fields: list[SchemaField]) -> dict:
    """Token-based methodology -> band routing from the Primary Methodology
    vocab. METHODOLOGY:-prefixed bands match by name token; otherwise any band
    matches by name token or by a field header containing every methodology
    token. 'market'/'transaction' tokens additionally route to the trading /
    transaction comps bands."""
    primary = next(
        (f for f in fields if f.header == _PRIMARY_METHODOLOGY_HEADER), None
    )
    if primary is None or not primary.controlled_vocab:
        raise ValueError(f"no controlled vocab compiled for {_PRIMARY_METHODOLOGY_HEADER!r}")

    band_first_col: dict[str, int] = {}
    for f in fields:
        band_first_col.setdefault(f.band, f.col_index)
    band_tokens = {band: set(normalize_text(band).split()) for band in band_first_col}
    header_tokens = [(f.band, set(normalize_text(f.header).split())) for f in fields]

    def _band_containing(marker: str) -> str:
        return next(band for band in band_first_col if marker in band)

    routing: dict[str, list[str]] = {}
    for methodology in primary.controlled_vocab:
        tokens = normalize_text(methodology).split()
        matched = {
            band
            for band in band_first_col
            if band.startswith(_METHODOLOGY_BAND_PREFIX)
            and any(tok in band_tokens[band] for tok in tokens)
        }
        if not matched:
            matched |= {
                band for band in band_first_col if any(tok in band_tokens[band] for tok in tokens)
            }
            matched |= {band for band, toks in header_tokens if set(tokens) <= toks}
        if "market" in tokens:
            matched.add(_band_containing(_TRADING_COMPS_MARKER))
        if "transaction" in tokens:
            matched.add(_band_containing(_TRANSACTION_COMPS_MARKER))
        routing[methodology] = sorted(matched, key=band_first_col.__getitem__)
    return routing


# --------------------------------------------------------------------------
# Entry point
# --------------------------------------------------------------------------


def _write_json(doc: dict, path: Path, pv_root: str) -> None:
    with guarded_open_write(path, pv_root) as fh:
        fh.write(json.dumps(doc, indent=2, ensure_ascii=False) + "\n")


def compile_schema(workbook_path: Path, out_dir: Path, pv_root: str) -> tuple[list[SchemaField], dict]:
    """Compile the workbook header rows into master_schema.json and
    band_routing.json under `out_dir`. Returns (fields, routing document).
    Output is byte-identical across recompiles of the same workbook."""
    workbook_path = Path(workbook_path)
    out_dir = Path(out_dir)
    bands, headers, descs = _read_header_rows(workbook_path)

    states = [_vocab_state(desc) for desc in descs]
    vocabs = _resolve_vocab_refs(states, headers, descs)
    fields = [
        _build_field(col, band, header, desc, vocab, state == "label")
        for col, (band, header, desc, vocab, (state, _)) in enumerate(
            zip(bands, headers, descs, vocabs, states, strict=True), start=1
        )
    ]

    schema_doc = {
        "source_workbook": f"{workbook_path.parent.name}/{workbook_path.name}",
        "sheet": SHEET_NAME,
        "field_count": len(fields),
        "fields": [f.model_dump() for f in fields],
    }
    routing_doc = {
        "derived_from": _ROUTING_DERIVED_FROM,
        "routing": _build_routing(fields),
    }
    _write_json(schema_doc, out_dir / MASTER_SCHEMA_FILENAME, pv_root)
    _write_json(routing_doc, out_dir / BAND_ROUTING_FILENAME, pv_root)

    dtype_counts: dict[str, int] = {}
    for f in fields:
        dtype_counts[f.dtype] = dtype_counts.get(f.dtype, 0) + 1
    log_event(
        logger,
        "schema compiled",
        workbook=str(workbook_path),
        out_dir=str(out_dir),
        field_count=len(fields),
        dtype_counts=dict(sorted(dtype_counts.items())),
        methodologies=list(routing_doc["routing"]),
    )
    return fields, routing_doc


if __name__ == "__main__":
    from pv_extractor.config import load_config
    from pv_extractor.logging_setup import setup_logging

    _project_root = Path(__file__).resolve().parents[3]
    _cfg = load_config(_project_root / "config.yaml")
    setup_logging(_cfg.output_dir, _cfg.pv_root, _cfg.logging.level)
    compile_schema(
        _project_root / "reference" / "master_index_v4.xlsx",
        _project_root / "schema",
        _cfg.pv_root,
    )
