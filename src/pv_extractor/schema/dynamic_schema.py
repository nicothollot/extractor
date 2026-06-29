"""Compile a schema from an ARBITRARY user-supplied reference workbook.

The legacy master template (``reference/master_index_v4.xlsx``) is a fixed
three-header-row sheet whose ~600 fields the deterministic band extractors are
hand-coded against (see ``schema/compile_schema.py``). An analyst, however, may
want to drop the tool's output into a completely different workbook: a new
sheet name, a single header row, and brand-new column headers. The deterministic
engine cannot read those headers (it only knows the master bands), so a custom
reference workbook is extracted **LLM-first** — each detected header becomes a
field the local LLM fills.

This module turns any workbook into ``(list[SchemaField], WorkbookLayout)``:

* the worksheet is autodetected (the sheet with the most populated header row,
  unless the workbook *is* the master, which is detected by an exact row-2
  header match and routed back to the committed master schema by the caller);
* the header row is autodetected (first row with >= 2 non-empty string cells);
* an optional description row directly beneath the header guides the LLM;
* required identity columns the pipeline emits (Memo ID, Run ID, Source
  Filename, Extraction Date, Valuation Date, Client, Deal) that are NOT already
  present are PREPENDED as the first columns — every other column keeps its
  order, shifted right by the number of prepended columns.

dtype inference reuses the master compiler's numeric/date/boolean cues so a
``Revenue ($M)`` header still parses as a USD-millions number.
"""

from __future__ import annotations

import logging
from pathlib import Path

from openpyxl import load_workbook

from pv_extractor.io_guard import open_read
from pv_extractor.models import SchemaField, WorkbookLayout
from pv_extractor.normalize import normalize_text
from pv_extractor.schema.compile_schema import (
    _DATE_HEADER_RE,
    _YN_RE,
    _numeric_dtype,
)

logger = logging.getLogger(__name__)

_ADMIN_BAND = "IDENTIFICATION"
_CUSTOM_BAND = "REFERENCE"
_MAX_HEADER_SCAN_ROWS = 12

# Canonical identity columns the run pipeline emits. Order is the prepend order.
# `header` is the column text written for a custom sheet; `aliases` are the
# normalized header forms that count as "already present" in the user's sheet
# (so we never duplicate a column the analyst already has). Matching is
# deliberately CONSERVATIVE — only near-exact names of obviously-administrative
# columns count, so a semantic user column like "Company" or "Manager" is left
# for the LLM to fill from the document rather than being stamped with the run's
# deal/client label. run.py writes each metadata value to the matching header.
ADMIN_COLUMNS: tuple[tuple[str, str, tuple[str, ...]], ...] = (
    ("memo_id", "Memo ID", ("memo id", "memoid")),
    ("run_id", "Run ID", ("run id", "runid")),
    ("source_filename", "Source Filename", ("source filename", "source file")),
    ("extraction_date", "Extraction Date", ("extraction date", "extracted on")),
    ("valuation_date", "Valuation Date", ("valuation date", "valuation as of")),
    ("client", "Client", ("client",)),
    ("deal", "Deal", ("deal",)),
)
_DATE_ROLES = frozenset({"extraction_date", "valuation_date"})


def _infer_dtype(header: str, desc: str) -> tuple[str, str | None]:
    """dtype/unit for a custom header, reusing the master numeric/date cues."""
    if _YN_RE.search(desc) or header.strip().endswith("Y/N"):
        return "boolean", None
    if _DATE_HEADER_RE.search(header):
        return "date", None
    numeric = _numeric_dtype(header)
    if numeric is not None:
        return numeric
    return "string", None


def _looks_like_descriptions(header_cells: list[str], next_cells: list[str]) -> bool:
    """True when the row below the header reads as descriptions/instructions
    (prose) rather than data — heuristic: more than half its non-empty cells
    are non-numeric and, on average, longer than the headers above them."""
    paired = [
        (h, n) for h, n in zip(header_cells, next_cells, strict=False) if h and n
    ]
    if len(paired) < 2:
        return False
    prose = 0
    longer = 0
    for h, n in paired:
        stripped = n.replace(",", "").replace(".", "").replace("%", "").strip()
        if not stripped.replace("-", "").isdigit():
            prose += 1
        if len(n) > len(h):
            longer += 1
    return prose >= len(paired) * 0.6 and longer >= len(paired) * 0.4


def _cell_str(value: object) -> str:
    return "" if value is None else str(value).strip()


def master_headers(schema_dir: Path) -> list[str]:
    """Row-2 headers of the committed master schema, for master detection."""
    import json

    path = Path(schema_dir) / "master_schema.json"
    with open_read(path) as fh:
        doc = json.load(fh)
    return [field["header"] for field in doc["fields"]]


def workbook_matches_master(workbook_path: Path, schema_dir: Path) -> bool:
    """True when this workbook's sheet "Index" row-2 headers exactly match the
    committed master schema — then the caller uses the fast deterministic path
    instead of the dynamic (LLM-first) one."""
    try:
        expected = master_headers(schema_dir)
    except (OSError, ValueError, KeyError):
        return False
    wb = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        if "Index" not in wb.sheetnames:
            return False
        ws = wb["Index"]
        actual = [
            _cell_str(c.value)
            for c in next(ws.iter_rows(min_row=2, max_row=2, max_col=len(expected)))
        ]
    except (StopIteration, KeyError):
        return False
    finally:
        wb.close()
    return actual == expected


def _pick_sheet(wb) -> str:
    """Sheet whose best candidate header row has the most non-empty string
    cells; ties resolve to the earliest sheet."""
    best_name = wb.sheetnames[0]
    best_score = -1
    for name in wb.sheetnames:
        ws = wb[name]
        for row in ws.iter_rows(min_row=1, max_row=_MAX_HEADER_SCAN_ROWS, values_only=True):
            score = sum(1 for v in row if isinstance(v, str) and v.strip())
            if score > best_score:
                best_score, best_name = score, name
            if score:
                break  # first populated row of this sheet is its header candidate
    return best_name


def _find_header_row(ws) -> int:
    """1-based index of the first row with >= 2 non-empty string cells."""
    for idx, row in enumerate(
        ws.iter_rows(min_row=1, max_row=_MAX_HEADER_SCAN_ROWS, values_only=True), start=1
    ):
        if sum(1 for v in row if isinstance(v, str) and v.strip()) >= 2:
            return idx
    return 1


def compile_schema_from_workbook(
    workbook_path: str | Path, *, sheet: str | None = None
) -> tuple[list[SchemaField], WorkbookLayout]:
    """Autodetect a custom reference workbook's fields and layout.

    Returns (schema_fields, layout). Identity columns the pipeline needs that
    are missing from the sheet are prepended as the first columns; the layout's
    ``prepended_admin`` lists them in order so the writer can insert them."""
    workbook_path = Path(workbook_path)
    wb = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        sheets = list(wb.sheetnames)
        sheet_name = sheet if sheet in sheets else _pick_sheet(wb)
        ws = wb[sheet_name]
        header_row = _find_header_row(ws)
        rows = list(
            ws.iter_rows(
                min_row=header_row, max_row=header_row + 1, values_only=True
            )
        )
        header_vals = list(rows[0]) if rows else []
        next_vals = list(rows[1]) if len(rows) > 1 else []
    finally:
        wb.close()

    # Trim trailing empty header columns.
    while header_vals and not _cell_str(header_vals[-1]):
        header_vals.pop()
    if not header_vals:
        raise ValueError(f"no header row found in {workbook_path.name!r} (sheet {sheet_name!r})")

    header_cells = [_cell_str(v) for v in header_vals]
    next_cells = [_cell_str(v) for v in next_vals]
    has_desc_row = _looks_like_descriptions(header_cells, next_cells)
    descs = next_cells if has_desc_row else [""] * len(header_cells)

    # User columns in original order (skip blank headers, keep position).
    user_cols: list[tuple[str, str]] = []  # (header, description)
    present_roles: dict[str, int] = {}  # role -> user position (0-based among user_cols)
    for header, desc in zip(header_cells, descs, strict=False):
        if not header:
            continue
        norm = normalize_text(header)
        pos = len(user_cols)
        for role, _admin_header, aliases in ADMIN_COLUMNS:
            if role in present_roles:
                continue
            if norm in aliases:
                present_roles[role] = pos
                break
        user_cols.append((header, desc))

    missing = [c for c in ADMIN_COLUMNS if c[0] not in present_roles]
    n_prepend = len(missing)

    fields: list[SchemaField] = []
    for col, (role, admin_header, _aliases) in enumerate(missing, start=1):
        dtype = "date" if role in _DATE_ROLES else "string"
        fields.append(
            SchemaField(
                col_index=col,
                band=_ADMIN_BAND,
                header=admin_header,
                description=f"Run identity column ({admin_header}).",
                dtype=dtype,
                required=role == "memo_id",
            )
        )
    for pos, (header, desc) in enumerate(user_cols):
        dtype, unit = _infer_dtype(header, desc)
        fields.append(
            SchemaField(
                col_index=n_prepend + pos + 1,
                band=_CUSTOM_BAND,
                header=header,
                description=desc or header,
                dtype=dtype,
                unit=unit,
            )
        )

    if "memo_id" in present_roles:
        memo_id_col = n_prepend + present_roles["memo_id"] + 1
    else:
        memo_id_col = next(f.col_index for f in fields if f.header == "Memo ID")

    # The OUTPUT is a fresh, clean sheet (write/workbook.create_custom_workbook) —
    # the user's own template is only the field definition and is never mutated.
    # So the layout describes that clean output: header row 1, data from row 2,
    # contiguous column indices. (The source's detected header/description rows
    # only fed field text above.)
    layout = WorkbookLayout(
        sheet_name=sheet_name,
        header_row=1,
        data_start_row=2,
        memo_id_col=memo_id_col,
        is_custom=True,
        prepended_admin=[c[1] for c in missing],
        sheets=sheets,
    )
    logger.info(
        "custom schema compiled: sheet=%r fields=%d prepended=%d desc_row=%s",
        sheet_name, len(fields), n_prepend, has_desc_row,
    )
    return fields, layout
