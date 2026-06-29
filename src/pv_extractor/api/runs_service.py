"""Run browsing for the dashboard/output screens.

A GUI run leaves a run_summary.json in its run directory; CLI runs do not,
so a fallback derives a minimal summary from the audit JSONs + cost ledger
(never from re-running anything). Review Flags / Run Log mirrors are read
straight from the run's own output workbook copy (read-only)."""

from __future__ import annotations

import json
import shutil
import zipfile
from datetime import datetime
from pathlib import Path

import openpyxl

from pv_extractor.config import Config
from pv_extractor.io_guard import guarded_open_write, open_read
from pv_extractor.llm.costs import LEDGER_FILENAME, read_ledger, summarize_ledger
from pv_extractor.write.workbook import FLAG_COLUMNS, RUN_LOG_COLUMNS

AUDIT_DIR = "audit"


def run_dirs(config: Config) -> list[Path]:
    output = Path(config.output_dir)
    if not output.exists():
        return []
    return sorted((p for p in output.glob("RUN_*") if p.is_dir()), reverse=True)


def load_audits(run_dir: Path) -> list[dict]:
    audits = []
    for path in sorted((run_dir / AUDIT_DIR).glob("*.json")):
        with open_read(path) as fh:
            audits.append(json.loads(fh.read().decode("utf-8")))
    return audits


def load_audit(run_dir: Path, memo_id: str) -> dict | None:
    path = run_dir / AUDIT_DIR / f"{memo_id}.json"
    if not path.exists():
        return None
    with open_read(path) as fh:
        return json.loads(fh.read().decode("utf-8"))


def _norm_path(p: str) -> str:
    return str(p or "").replace("\\", "/").strip().lower()


def find_runs_for_file(config: Config, file_path: str) -> list[dict]:
    """Prior runs that already extracted this exact document (matched by file
    path, newest first). Powers the Direct Run 'use previous output or re-run?'
    prompt so a cached extraction is never silently reused."""
    target = _norm_path(file_path)
    if not target:
        return []
    matches: list[dict] = []
    for run_dir in run_dirs(config):
        for audit in load_audits(run_dir):
            if _norm_path(audit.get("file_path", "")) == target:
                _, finished, _ = _run_timing(run_dir)
                matches.append({
                    "run_id": run_dir.name,
                    "memo_id": audit.get("memo_id"),
                    "finished_at": finished,
                    "has_workbook": workbook_path(run_dir) is not None,
                })
                break  # one hit per run is enough
    return matches


def file_shas_for_run(run_dir: Path) -> list[str]:
    return [s for a in load_audits(run_dir) if (s := a.get("file_sha256"))]


def delete_run(config: Config, run_id: str) -> dict:
    """Remove a run from history: delete its output directory and forget every
    cached extraction for the files it contained, so re-running re-extracts.
    Only deletes inside output_dir under a validated RUN_ id."""
    if "/" in run_id or "\\" in run_id or not run_id.startswith("RUN_"):
        raise ValueError(f"invalid run id {run_id!r}")
    run_dir = Path(config.output_dir) / run_id
    if not run_dir.is_dir():
        raise FileNotFoundError(run_id)
    shas = file_shas_for_run(run_dir)
    forgotten = 0
    try:
        from pv_extractor.extract import cache as result_cache
        from pv_extractor.indexer import db

        conn = db.open_db(config.db_path, config.pv_root)
        try:
            result_cache.init_cache(conn)
            forgotten = result_cache.forget_by_sha256(conn, shas)
        finally:
            conn.close()
    except Exception:  # noqa: BLE001 — forgetting the cache is best-effort
        forgotten = 0
    shutil.rmtree(run_dir)
    return {"run_id": run_id, "cache_entries_forgotten": forgotten}


def ledger_summary(run_dir: Path) -> dict | None:
    path = run_dir / "llm" / LEDGER_FILENAME
    if not path.exists():
        return None
    return summarize_ledger(read_ledger(path))


def ledger_entries(run_dir: Path) -> list[dict]:
    path = run_dir / "llm" / LEDGER_FILENAME
    return read_ledger(path) if path.exists() else []


def run_diagnostics(run_dir: Path) -> dict:
    path = run_dir / "diagnostics.json"
    if not path.exists():
        return {}
    with open_read(path) as fh:
        data = json.loads(fh.read().decode("utf-8"))
    diagnostics = data.get("diagnostics")
    return diagnostics if isinstance(diagnostics, dict) else {}


def workbook_path(run_dir: Path) -> Path | None:
    matches = sorted(run_dir.glob("master_index_*.xlsx"))
    return matches[0] if matches else None


def run_summary(run_dir: Path) -> dict:
    """run_summary.json when present (GUI run), else derived from audits."""
    summary_path = run_dir / "run_summary.json"
    if summary_path.exists():
        with open_read(summary_path) as fh:
            return json.loads(fh.read().decode("utf-8"))
    return _derive_summary(run_dir)


def _run_timing(run_dir: Path) -> tuple[str | None, str | None, float | None]:
    """(started_at, finished_at, duration_minutes) for a CLI run that left no
    run_summary.json: the start time is encoded in the run id (RUN_YYYYMMDD_
    HHMMSS); the finish is the newest audit/workbook mtime (last thing written)."""
    started: str | None = None
    try:
        started_dt = datetime.strptime(run_dir.name, "RUN_%Y%m%d_%H%M%S")
        started = started_dt.isoformat(timespec="seconds")
    except ValueError:
        started_dt = None
    newest = 0.0
    for path in [*(run_dir / AUDIT_DIR).glob("*.json"), *run_dir.glob("master_index_*.xlsx")]:
        try:
            newest = max(newest, path.stat().st_mtime)
        except OSError:
            continue
    finished: str | None = None
    duration: float | None = None
    if newest > 0:
        finished_dt = datetime.fromtimestamp(newest)
        finished = finished_dt.isoformat(timespec="seconds")
        if started_dt is not None:
            duration = round(max(0.0, (finished_dt - started_dt).total_seconds()) / 60.0, 2)
    return started, finished, duration


def _derive_summary(run_dir: Path) -> dict:
    audits = load_audits(run_dir)
    qa_counts = {"qa_pass": 0, "qa_pass_with_flags": 0, "qa_fail": 0}
    flags = 0
    clients: set[str] = set()
    deals: set[str] = set()
    companies: set[str] = set()
    source_files = 0
    for audit in audits:
        if audit.get("client"):
            clients.add(audit["client"])
        if audit.get("deal"):
            deals.add(audit["deal"])
        if audit.get("file_path"):
            source_files += 1
        for asset in audit.get("assets", []):
            status = asset.get("qa_status", "")
            if status in qa_counts:
                qa_counts[status] += 1
            flags += len(asset.get("flags", []))
            if asset.get("asset_name"):
                companies.add(str(asset["asset_name"]))
    ledger = ledger_summary(run_dir)
    wb = workbook_path(run_dir)
    started_at, finished_at, duration_minutes = _run_timing(run_dir)
    return {
        "run_id": run_dir.name,
        "source": "cli",
        "started_at": started_at,
        "finished_at": finished_at,
        "dry_run": False,
        "scope": None,
        "period": None,
        "coverage": [],
        "coverage_counts": {},
        "clients": sorted(c for c in clients if c),
        "deals": sorted(deals),
        "companies": sorted(companies),
        "source_files": source_files,
        "sources": [
            {
                "file_name": a.get("file_name"), "file_path": a.get("file_path"),
                "client": a.get("client"), "deal": a.get("deal"),
            }
            for a in audits if a.get("file_path")
        ],
        "memos": len(audits),
        "assets": sum(len(a.get("assets", [])) for a in audits),
        "rows_added": None,
        "flags_added": flags,
        "qa_counts": qa_counts,
        "duration_minutes": duration_minutes,
        "workbook_path": str(wb) if wb else None,
        "llm": {
            "enabled": ledger is not None,
            "executed": ledger is not None,
            "attempts": ledger["attempts"] if ledger else 0,
            "cache_hits": ledger["cache_hits"] if ledger else 0,
            "memos_escalated": ledger["memos"] if ledger else 0,
            "memos_deferred": 0,
            "total_cost_usd": ledger["total_usd"] if ledger else 0.0,
            "cost_source": (
                "actual+estimated" if ledger and ledger["actual_usd"] > 0 else
                ("estimated" if ledger else None)
            ),
            "detail": "",
        },
        "diagnostics": run_diagnostics(run_dir),
    }


def list_run_summaries(config: Config, limit: int = 50) -> list[dict]:
    return [run_summary(run_dir) for run_dir in run_dirs(config)[:limit]]


def cost_history(config: Config) -> list[dict]:
    """Per-run LLM spend for the dashboard sparkline (oldest first)."""
    points = []
    for run_dir in reversed(run_dirs(config)):
        ledger = ledger_summary(run_dir)
        points.append(
            {
                "run_id": run_dir.name,
                "total_usd": ledger["total_usd"] if ledger else 0.0,
                "actual_usd": ledger["actual_usd"] if ledger else 0.0,
            }
        )
    return points


def _sheet_rows(path: Path, sheet_name: str, columns: tuple[str, ...], skip_header_rows: int) -> list[dict]:
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook[sheet_name]
        rows = []
        for row in sheet.iter_rows(min_row=skip_header_rows + 1, values_only=True):
            if all(cell is None for cell in row):
                continue
            rows.append({col: row[i] if i < len(row) else None for i, col in enumerate(columns)})
        return rows
    finally:
        workbook.close()


def review_flags_mirror(run_dir: Path) -> list[dict]:
    wb = workbook_path(run_dir)
    return _sheet_rows(wb, "Review Flags", FLAG_COLUMNS, 1) if wb else []


def run_log_mirror(run_dir: Path) -> list[dict]:
    wb = workbook_path(run_dir)
    return _sheet_rows(wb, "Run Log", RUN_LOG_COLUMNS, 1) if wb else []


# Compact preview of the key extracted columns per memo/asset for the per-run
# Output page — read straight from the run's own Index sheet (read-only), keyed
# by clean field names so the typed frontend never depends on the emoji header.
_INDEX_PREVIEW_COLUMNS: tuple[tuple[str, str], ...] = (
    ("memo_id", "\U0001f511 Memo ID"),
    ("fund_manager", "Fund Manager"),
    ("portfolio_company", "Portfolio Company"),
    ("reporting_period", "Reporting Period"),
    ("valuation_date", "Valuation Date"),
    ("primary_methodology", "Primary Methodology"),
    ("headline_value", "Implied Equity Value 100% ($M)"),
    ("moic", "MOIC"),
)


def _qa_status_by_row_memo(run_dir: Path) -> dict[str, str]:
    """row_memo_id -> qa_status from the run's audit JSONs (QA verdicts live in
    the audits / Review Flags sheet, not as an Index column)."""
    by_row: dict[str, str] = {}
    for audit in load_audits(run_dir):
        for asset in audit.get("assets", []):
            row_id = asset.get("row_memo_id")
            if row_id:
                by_row[str(row_id)] = asset.get("qa_status", "")
    return by_row


def index_rows_mirror(run_dir: Path, run_id: str | None = None) -> list[dict]:
    """Curated subset of this run's Index rows (company, period, methodology,
    headline valuation, MOIC) plus the per-row QA status merged from the
    audits. Filtered to `run_id` so a cumulative workbook still shows only the
    rows this run added. Read-only — never re-extracts anything."""
    wb = workbook_path(run_dir)
    if wb is None:
        return []
    qa_by_row = _qa_status_by_row_memo(run_dir)
    # A custom-reference run wrote to its own sheet at its own header/data rows;
    # the snapshot records that layout. Master runs leave no snapshot -> the
    # committed master layout (sheet "Index", header row 2, data row 4).
    sheet_name, header_row_idx, data_start_row, memo_header = "Index", 2, 4, "\U0001f511 Memo ID"
    snapshot = run_dir / "schema_snapshot.json"
    if snapshot.is_file():
        try:
            doc = json.loads(snapshot.read_text(encoding="utf-8"))
            layout = doc.get("layout") or {}
            sheet_name = layout.get("sheet_name", sheet_name)
            header_row_idx = layout.get("header_row", header_row_idx)
            data_start_row = layout.get("data_start_row", data_start_row)
            memo_header = "Memo ID"
        except (OSError, ValueError):
            pass
    workbook = openpyxl.load_workbook(wb, read_only=True, data_only=True)
    try:
        if sheet_name not in workbook.sheetnames:
            return []
        sheet = workbook[sheet_name]
        header_row = next(
            sheet.iter_rows(min_row=header_row_idx, max_row=header_row_idx, values_only=True), ()
        )
        col_by_header = {h: i for i, h in enumerate(header_row) if h is not None}
        run_col = col_by_header.get("Run ID")
        memo_col = col_by_header.get(memo_header)
        rows: list[dict] = []
        for row in sheet.iter_rows(min_row=data_start_row, values_only=True):
            if all(cell is None for cell in row):
                continue
            if run_id is not None and run_col is not None:
                value = row[run_col] if run_col < len(row) else None
                if value != run_id:
                    continue
            entry: dict = {}
            for key, header in _INDEX_PREVIEW_COLUMNS:
                idx = col_by_header.get(header)
                entry[key] = row[idx] if idx is not None and idx < len(row) else None
            memo_id = row[memo_col] if memo_col is not None and memo_col < len(row) else None
            entry["qa_status"] = qa_by_row.get(str(memo_id), "") if memo_id is not None else ""
            rows.append(entry)
        return rows
    finally:
        workbook.close()


def build_audit_zip(run_dir: Path, config: Config) -> Path:
    """Zip the run's audit JSONs for download (rebuilt when stale)."""
    zip_path = run_dir / "gui" / "audits.zip"
    audit_files = sorted((run_dir / AUDIT_DIR).glob("*.json"))
    if zip_path.exists() and audit_files:
        newest = max(p.stat().st_mtime for p in audit_files)
        if zip_path.stat().st_mtime >= newest:
            return zip_path
    with guarded_open_write(zip_path, config.pv_root, mode="wb") as fh:
        with zipfile.ZipFile(fh, "w", compression=zipfile.ZIP_DEFLATED) as archive:
            for path in audit_files:
                with open_read(path) as src:
                    archive.writestr(f"audit/{path.name}", src.read())
    return zip_path
