"""Dynamic reference workbooks + Direct Run.

Covers the autodetecting schema compiler (any workbook -> fields + layout,
identity columns prepended), the layout-driven writer adapting to a custom
sheet, and the Direct Run path (extract ONE explicit file, no locator) against
both the master template and a custom reference workbook (LLM-first, fake CLI).
"""

from __future__ import annotations

from datetime import datetime
from pathlib import Path

import openpyxl
import pytest
from fixtures.docgen import make_text_pdf
from fixtures.fake_claude import FakeClaudeCodeClient, field_result

from pv_extractor.llm.escalate import resolve_settings
from pv_extractor.models import FieldHit
from pv_extractor.run import run
from pv_extractor.schema.dynamic_schema import (
    compile_schema_from_workbook,
    workbook_matches_master,
)
from pv_extractor.write.workbook import WorkbookWriter, create_custom_workbook

PROJECT_ROOT = Path(__file__).resolve().parent.parent
MASTER_TEMPLATE = PROJECT_ROOT / "reference" / "master_index_v4.xlsx"


def _custom_workbook(path: Path, *, sheet="Valuations", headers=None, desc_row=False) -> Path:
    headers = headers or ["Company", "NAV ($M)", "IRR %", "Notes"]
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet
    ws.append(headers)
    if desc_row:
        ws.append([f"description for {h}" for h in headers])
    wb.save(path)
    return path


# ---------------------------------------------------------------------------
# dynamic schema compiler
# ---------------------------------------------------------------------------


def test_master_template_is_detected_as_master():
    assert workbook_matches_master(MASTER_TEMPLATE, PROJECT_ROOT / "schema") is True


def test_custom_workbook_autodetects_fields_and_prepends_identity(tmp_path):
    src = _custom_workbook(tmp_path / "custom.xlsx")
    assert workbook_matches_master(src, PROJECT_ROOT / "schema") is False

    fields, layout = compile_schema_from_workbook(src)
    assert layout.is_custom and layout.sheet_name == "Valuations"
    assert layout.header_row == 1 and layout.data_start_row == 2
    # Memo ID is prepended at column 1 (the writer/dedupe key).
    assert layout.memo_id_col == 1
    assert fields[0].header == "Memo ID" and fields[0].band == "IDENTIFICATION"
    headers = [f.header for f in fields]
    assert "Memo ID" in layout.prepended_admin
    # User columns survive, shifted right by the prepended block, dtype inferred.
    assert headers[-4:] == ["Company", "NAV ($M)", "IRR %", "Notes"]
    nav = next(f for f in fields if f.header == "NAV ($M)")
    assert nav.dtype == "number" and nav.band == "REFERENCE"
    irr = next(f for f in fields if f.header == "IRR %")
    assert irr.dtype == "percent"


def test_description_row_feeds_field_text(tmp_path):
    """A description row beneath the headers guides the LLM (becomes field.description);
    the clean OUTPUT layout is still header row 1 / data row 2."""
    src = _custom_workbook(tmp_path / "desc.xlsx", desc_row=True)
    fields, layout = compile_schema_from_workbook(src)
    assert layout.header_row == 1 and layout.data_start_row == 2
    company = next(f for f in fields if f.header == "Company")
    assert company.description == "description for Company"


# ---------------------------------------------------------------------------
# layout-driven writer on a custom sheet
# ---------------------------------------------------------------------------


def test_clean_workbook_writes_by_index(tmp_path):
    """A custom reference yields a FRESH clean sheet (the user's own template is
    never mutated); headers and values align by schema column index."""
    src = _custom_workbook(tmp_path / "custom.xlsx")
    fields, layout = compile_schema_from_workbook(src)
    fake_pv = str(tmp_path / "pv")
    dest = tmp_path / "out.xlsx"
    create_custom_workbook(dest, fields, layout, fake_pv)
    writer = WorkbookWriter(dest, fields, fake_pv, layout=layout)
    by_h = {f.header: f for f in fields}
    writer.append_index_row([
        FieldHit(field="Memo ID", col_index=by_h["Memo ID"].col_index, band="IDENTIFICATION",
                 value="MEMO_X_001", method="metadata", confidence=1.0, evidence="x"),
        FieldHit(field="Company", col_index=by_h["Company"].col_index, band="REFERENCE",
                 value="Acme Corp", method="llm:claude", confidence=0.9, evidence="x"),
    ])
    writer.save()

    out = openpyxl.load_workbook(dest)["Valuations"]
    assert out.cell(row=1, column=1).value == "Memo ID"  # prepended at the front
    company_col = by_h["Company"].col_index
    # value lands UNDER its own header — the misalignment bug regression check
    assert out.cell(row=1, column=company_col).value == "Company"
    assert out.cell(row=2, column=1).value == "MEMO_X_001"
    assert out.cell(row=2, column=company_col).value == "Acme Corp"
    assert writer.find_index_row("MEMO_X_001") == 2
    assert "Review Flags" in writer.workbook.sheetnames
    assert "Run Log" in writer.workbook.sheetnames


# ---------------------------------------------------------------------------
# Direct Run
# ---------------------------------------------------------------------------


def _memo_pdf(path: Path) -> Path:
    make_text_pdf(path, [[
        "Valuation Memo",
        "Fund Name: Direct Fund I",
        "Gross IRR: 18.0%",
        "MOIC: 2.1x",
    ]])
    return path


def test_direct_run_master_template_extracts_one_file(phase2_env, tmp_path):
    doc = _memo_pdf(tmp_path / "memo.pdf")
    fake = FakeClaudeCodeClient({
        "Gross IRR %": field_result(18.0, unit="percent", page=1, quote="Gross IRR: 18.0%"),
    })
    report = run(
        phase2_env, scope="deal", period="", direct_file=str(doc),
        direct_client="Acme Capital", direct_deal="Project Atlas",
        now=datetime(2026, 6, 25, 9, 0, 0),
        llm_settings=resolve_settings(phase2_env), llm_client=fake,
    )
    assert report.coverage and report.coverage[0].status == "FOUND"
    assert report.rows_added >= 1
    memo = report.memos[0]
    assert memo.client == "Acme Capital" and memo.deal == "Project Atlas"
    assert memo.file_path == str(doc)
    # no schema snapshot for the master path (review falls back to committed schema)
    assert not (report.run_dir / "schema_snapshot.json").exists()


def test_find_and_delete_run_forgets_history(phase2_env, tmp_path):
    """A run is discoverable by its source file, and deleting it removes the
    output dir AND forgets the cached extraction (so it re-extracts later)."""
    from pv_extractor.api import runs_service

    doc = _memo_pdf(tmp_path / "memo.pdf")
    fake = FakeClaudeCodeClient({
        "Gross IRR %": field_result(18.0, unit="percent", page=1, quote="Gross IRR: 18.0%"),
    })
    report = run(
        phase2_env, scope="deal", period="", direct_file=str(doc),
        now=datetime(2026, 6, 25, 12, 0, 0),
        llm_settings=resolve_settings(phase2_env), llm_client=fake,
    )
    run_id = report.run_id

    found = runs_service.find_runs_for_file(phase2_env, str(doc))
    assert any(m["run_id"] == run_id for m in found)

    runs_service.delete_run(phase2_env, run_id)
    assert not report.run_dir.exists()
    assert not any(m["run_id"] == run_id for m in runs_service.find_runs_for_file(phase2_env, str(doc)))


def test_direct_run_custom_reference_is_llm_first(phase2_env, tmp_path):
    doc = _memo_pdf(tmp_path / "memo.pdf")
    template = _custom_workbook(tmp_path / "custom.xlsx")
    fake = FakeClaudeCodeClient({
        "Company": field_result("Direct Fund I", page=1, quote="Fund Name: Direct Fund I"),
        "IRR %": field_result(18.0, unit="percent", page=1, quote="Gross IRR: 18.0%"),
    })
    report = run(
        phase2_env, scope="deal", period="", direct_file=str(doc),
        direct_client="Acme Capital", direct_deal="Project Atlas",
        template=str(template), now=datetime(2026, 6, 25, 10, 0, 0),
        llm_settings=resolve_settings(phase2_env), llm_client=fake,
    )
    assert report.coverage[0].status == "FOUND"
    # a custom reference persists its schema so the review queue can reopen it
    assert (report.run_dir / "schema_snapshot.json").is_file()

    out = openpyxl.load_workbook(report.workbook_path)["Valuations"]
    headers = [out.cell(row=1, column=c).value for c in range(1, 12)]
    assert "Memo ID" in headers and "Company" in headers
    memo_col = headers.index("Memo ID") + 1
    company_col = headers.index("Company") + 1
    # the one extracted row carries the prepended Memo ID + the LLM-filled value,
    # written UNDER its own header (column-alignment regression check)
    row = next(
        r for r in range(2, out.max_row + 1)
        if out.cell(row=r, column=memo_col).value == report.memos[0].memo_id
    )
    assert out.cell(row=row, column=company_col).value == "Direct Fund I"

    # confidence_selection is OFF for LLM-first runs: the review shows the
    # MODEL's confidence (field_result default "high" -> 0.85), never the
    # ungrounded 0.20 cap.
    company_hit = next(h for h in report.memos[0].assets[0].hits if h.field == "Company")
    assert company_hit.confidence > 0.5
    assert company_hit.method.startswith("llm:")


def test_custom_reference_without_llm_is_refused(phase2_env, tmp_path):
    doc = _memo_pdf(tmp_path / "memo.pdf")
    template = _custom_workbook(tmp_path / "custom.xlsx")
    with pytest.raises(ValueError, match="custom reference workbook"):
        run(
            phase2_env, scope="deal", period="", direct_file=str(doc),
            template=str(template), now=datetime(2026, 6, 25, 11, 0, 0),
        )  # llm_settings=None -> LLM disabled
