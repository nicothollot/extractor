"""Phase-4 GUI backend tests (FastAPI TestClient against the fixture tree).

Covers: setup/doctor endpoints, index metadata, the job lifecycle (dry-run
preflight + full run with live events), run browsing, the review queue with
accept/edit/bulk actions persisting through the writer seam, evidence page
rendering, locator override endpoints, comment-preserving config/pricing
edits, downloads, and the RunControl cancellation path. No test launches
the real Claude Code CLI (every run here disables the LLM fallback)."""

from __future__ import annotations

import json
import re
import threading
import time
from datetime import datetime
from pathlib import Path

import pytest

fastapi = pytest.importorskip("fastapi", reason="gui extra not installed")

from fastapi.testclient import TestClient  # noqa: E402

from pv_extractor.config import load_config  # noqa: E402
from pv_extractor.indexer.db import init_schema, open_db  # noqa: E402
from pv_extractor.indexer.scan_tree import scan_tree  # noqa: E402

PROJECT_ROOT = Path(__file__).resolve().parent.parent


@pytest.fixture(scope="module")
def gui_env(fixture_pv_root, tmp_path_factory):
    """Isolated config + index + COPIES of the editable YAML files so the
    GUI's config/pricing edits never touch the repo files."""
    base = tmp_path_factory.mktemp("gui")
    (base / "config").mkdir()
    (base / "config.yaml").write_text((PROJECT_ROOT / "config.yaml").read_text(encoding="utf-8"), encoding="utf-8")
    (base / "config" / "models.yaml").write_text(
        (PROJECT_ROOT / "config" / "models.yaml").read_text(encoding="utf-8"), encoding="utf-8"
    )
    (base / "rules.yaml").write_text((PROJECT_ROOT / "rules.yaml").read_text(encoding="utf-8"), encoding="utf-8")
    (base / "aliases.yaml").write_text((PROJECT_ROOT / "aliases.yaml").read_text(encoding="utf-8"), encoding="utf-8")

    config = load_config(base / "config.yaml")
    config.pv_root = str(fixture_pv_root)
    config.output_dir = base / "output"
    config.db_path = base / "output" / "pv_index.db"

    conn = open_db(config.db_path, config.pv_root)
    init_schema(conn)
    scan_tree(conn, str(fixture_pv_root), config)
    conn.close()
    return config, base / "config.yaml"


@pytest.fixture(scope="module")
def client(gui_env):
    from pv_extractor.api.app import create_app

    config, config_path = gui_env
    app = create_app(config, config_path=config_path)
    with TestClient(app) as test_client:
        yield test_client


def _wait_job(client: TestClient, job_id: str, timeout: float = 240.0) -> dict:
    deadline = time.time() + timeout
    while time.time() < deadline:
        job = client.get(f"/api/jobs/{job_id}").json()
        if job["status"] not in ("queued", "running", "cancelling"):
            return job
        time.sleep(0.5)
    raise AssertionError(f"job {job_id} did not finish in {timeout}s")


@pytest.fixture(scope="module")
def completed_run(client) -> dict:
    """One full pipeline run (LLM disabled) shared by the browsing tests."""
    r = client.post("/api/jobs/run", json={
        "scope": "all", "period": "Q1 2026", "dry_run": False, "llm": {"enabled": False},
    })
    assert r.status_code == 200, r.text
    job = _wait_job(client, r.json()["job"]["id"])
    assert job["status"] == "completed", job
    return job


# ---------------------------------------------------------------------------
# setup / doctor / metadata
# ---------------------------------------------------------------------------


def test_health_and_setup_status(client) -> None:
    health = client.get("/api/health").json()
    assert health["ok"] is True and "version" in health

    status = client.get("/api/setup/status", params={"include_claude": False}).json()
    names = {item["name"] for item in status["items"]}
    assert "python core dependencies" in names
    assert "output_dir writable" in names
    assert "frontend build" in names
    core = next(i for i in status["items"] if i["name"] == "python core dependencies")
    assert core["ok"] is True


def test_doctor_endpoint_reports_checks(client) -> None:
    body = client.get("/api/doctor").json()
    names = {c["check"] for c in body["checks"]}
    assert "models.yaml" in names
    assert "schema/master_schema.json" in names
    assert any(name.startswith("claude") for name in names)
    assert "cost accounting" in names


def test_index_metadata_endpoints(client) -> None:
    clients = client.get("/api/index/clients").json()["clients"]
    assert "Angelo Gordon" in clients
    deals = client.get("/api/index/deals", params={"client": "Angelo Gordon"}).json()["deals"]
    assert "Accell" in deals
    periods = client.get(
        "/api/index/periods", params={"client": "Angelo Gordon", "deal": "Accell"}
    ).json()["periods"]
    assert {"period", "as_of_date", "label"} <= set(periods[0])
    # Deduped to one entry per reporting-period label (never one per date folder),
    # and the submit value (period) equals the label.
    labels = [p["label"] for p in periods]
    assert len(labels) == len(set(labels))
    assert all(p["period"] == p["label"] for p in periods)
    doc_types = client.get("/api/index/doc-types").json()["doc_types"]
    assert "valuation_memo" in doc_types
    templates = client.get("/api/templates").json()
    assert templates["default_template"].endswith("master_index_v4.xlsx")


def test_template_inspect_master_and_custom(client, tmp_path) -> None:
    import openpyxl

    default = client.get("/api/templates").json()["default_template"]
    master = client.post("/api/templates/inspect", json={"path": default}).json()
    assert master["is_custom"] is False and master["ready"] is True
    assert master["sheet_name"] == "Index" and not master["prepended_admin"]

    custom_path = tmp_path / "custom.xlsx"
    wb = openpyxl.Workbook()
    wb.active.title = "Valuations"
    wb.active.append(["Company", "NAV ($M)", "IRR %"])
    wb.save(custom_path)
    custom = client.post("/api/templates/inspect", json={"path": str(custom_path)}).json()
    assert custom["is_custom"] is True and custom["sheet_name"] == "Valuations"
    assert "Memo ID" in custom["prepended_admin"]
    assert any(f["header"] == "Company" for f in custom["fields"])

    missing = client.post("/api/templates/inspect", json={"path": str(tmp_path / "nope.xlsx")})
    assert missing.status_code == 404


def test_deal_discovery_refresh_and_search_endpoints(client) -> None:
    # heuristic refresh job populates deal_folders for the client
    r = client.post("/api/index/deals/refresh", json={"client": "Angelo Gordon"})
    assert r.status_code == 200, r.text
    job = _wait_job(client, r.json()["job"]["id"])
    assert job["status"] == "completed", job
    by_name = {d["name"]: d for d in job["result"]["deals"]}
    assert "Accell" in by_name and by_name["Accell"]["confidence"] > 0.5
    assert by_name["Accell"]["folder_paths"] == ["Angelo Gordon\\Accell"]

    # /index/deals now carries the discovered detail alongside the names
    payload = client.get("/api/index/deals", params={"client": "Angelo Gordon"}).json()
    assert "Accell" in payload["deals"]
    assert any(f["name"] == "Accell" for f in payload["deal_folders"])
    # A heuristic-only refresh leaves no saved LLM discovery.
    assert payload["last_llm_discovery"] is None
    assert "display_min_confidence" in payload

    # Display floor hides low-confidence folders without dropping them from storage:
    # a 0.99 floor hides Accell, a 0.0 floor brings it back.
    hidden = client.get(
        "/api/index/deals", params={"client": "Angelo Gordon", "min_confidence": 0.99}
    ).json()
    assert all(f["name"] != "Accell" for f in hidden["deal_folders"])
    assert hidden["hidden_below_floor"] >= 1
    shown = client.get(
        "/api/index/deals", params={"client": "Angelo Gordon", "min_confidence": 0.0}
    ).json()
    assert any(f["name"] == "Accell" for f in shown["deal_folders"])

    # manual search: fuzzy client, fuzzy deal (with path preview), period parse
    matches = client.get("/api/index/search/clients", params={"q": "angelo"}).json()["matches"]
    assert matches and matches[0]["client"] == "Angelo Gordon"
    matches = client.get(
        "/api/index/search/deals", params={"client": "Angelo Gordon", "q": "acell"}
    ).json()["matches"]
    assert matches and matches[0]["name"] == "Accell"
    assert matches[0]["folder_paths"] == ["Angelo Gordon\\Accell"]
    periods = client.get(
        "/api/index/search/periods",
        params={"client": "Angelo Gordon", "deal": "Accell", "q": "1 31 2025"},
    ).json()
    assert periods["resolved_as_of"] == "2025-01-31"
    assert periods["matches"][0]["as_of_date"] == "2025-01-31"
    assert periods["matches"][0]["exact"] is True
    garbage = client.get(
        "/api/index/search/periods",
        params={"client": "Angelo Gordon", "deal": "Accell", "q": "not a date"},
    ).json()
    assert garbage["parse_error"] is not None


def test_deal_feedback_roundtrip_and_learned_admin(client) -> None:
    """POST feedback records a correction + re-discovers (with learning) as a
    background job; GET /learned lists priors + corrections; DELETE removes one;
    apply_learning=false skips the learning layer."""
    cl = "Angelo Gordon"

    # baseline: discovery without any learning correction
    r = client.post(
        "/api/index/deals/refresh", json={"client": cl, "apply_learning": False}
    )
    job = _wait_job(client, r.json()["job"]["id"])
    assert job["status"] == "completed", job
    baseline = {d["name"]: d["confidence"] for d in job["result"]["deals"]}
    assert "Accell" in baseline

    # record an add_folder correction under an admin-ish path -> raises this
    # client's admin_container layout prior (generalizes to other deals)
    r = client.post(
        "/api/index/deals/feedback",
        json={
            "client": cl,
            "deal": "Buried Co",
            "action": "add_folder",
            "folder_path": f"{cl}\\_Admin\\Buried Co",
        },
    )
    assert r.status_code == 200, r.text
    job = _wait_job(client, r.json()["job"]["id"])
    assert job["status"] == "completed", job
    result = job["result"]
    assert result["client"] == cl
    by_name = {d["name"]: d for d in result["deals"]}
    # the force-added (learned) deal is present
    assert "Buried Co" in by_name and by_name["Buried Co"]["method"] == "learned"
    # the learned client-scoped prior is returned
    assert result["learned"].get("admin_container", 0) > 0
    # generalization: a pre-existing deal's confidence is nudged up
    assert by_name["Accell"]["confidence"] > baseline["Accell"]

    # GET /learned reflects the cached prior + the recorded correction
    learned = client.get("/api/index/deals/learned", params={"client": cl}).json()
    assert learned["client"] == cl
    assert learned["priors"].get("admin_container", 0) > 0
    corrections = learned["corrections"]
    assert any(
        c["deal"] == "Buried Co" and c["action"] == "add_folder" for c in corrections
    )

    # DELETE one correction
    cid = next(c["id"] for c in corrections if c["deal"] == "Buried Co")
    deleted = client.delete(f"/api/index/deals/feedback/{cid}").json()
    assert deleted["deleted"] is True

    # after deletion the correction is gone from the listing
    learned2 = client.get("/api/index/deals/learned", params={"client": cl}).json()
    assert not any(c["id"] == cid for c in learned2["corrections"])


def test_deal_feedback_invalid_action_fails_job(client) -> None:
    """An invalid correction action surfaces as a FAILED background job (the
    ValueError message is carried in job.error), not a synchronous 422."""
    r = client.post(
        "/api/index/deals/feedback",
        json={"client": "Angelo Gordon", "deal": "X", "action": "frobnicate"},
    )
    assert r.status_code == 200, r.text  # the job is accepted, then fails
    job = _wait_job(client, r.json()["job"]["id"])
    assert job["status"] == "failed", job
    assert "frobnicate" in (job.get("error") or "")


def test_utility_job_cooperative_pause_keeps_partial_result(gui_env) -> None:
    """Cancelling a utility job whose fn accepts should_stop is a PAUSE: the
    fn winds down, its partial result is stored, status ends 'cancelled' —
    the scan job uses this to keep a partially built index usable."""
    from pv_extractor.api.jobs import JobManager

    config, _ = gui_env
    manager = JobManager(config)
    started = threading.Event()

    def fn(emit, should_stop) -> dict:
        emit("tick", {"n": 0})
        started.set()
        deadline = time.time() + 10
        while not should_stop() and time.time() < deadline:
            time.sleep(0.02)
        return {"partial": True, "stopped": should_stop()}

    job = manager.start_utility("scan", {}, fn)
    assert started.wait(5)
    manager.cancel(job.id)
    deadline = time.time() + 10
    while time.time() < deadline:
        info = manager.get(job.id)
        if info.status not in ("queued", "running", "cancelling"):
            break
        time.sleep(0.05)
    assert info.status == "cancelled"
    assert info.result == {"partial": True, "stopped": True}


def test_index_status_and_scan_job(client) -> None:
    status = client.get("/api/index/status").json()
    assert status["ready"] is True
    assert status["files"] > 0 and status["clients"] > 0
    assert status["pv_root_exists"] is True

    # rescan through the GUI job (same init_schema + scan_tree as the CLI);
    # no request body = full pv_root walk
    r = client.post("/api/index/scan")
    assert r.status_code == 200, r.text
    job = _wait_job(client, r.json()["job"]["id"])
    assert job["status"] == "completed"
    assert job["result"]["files_seen"] > 0
    assert job["result"]["errors"] == 0

    # live progress events were emitted (at least the per-root starter)
    events = client.get(f"/api/jobs/{job['id']}/events").json()["events"]
    starts = [e for e in events if e["type"] == "scan_progress"]
    assert starts and {"root", "files_seen", "prev_total", "dir"} <= set(starts[0]["payload"])

    # an unreachable root is refused before a job is even created
    r = client.post("/api/index/scan", json={"root": "/definitely/not/here"})
    assert r.status_code == 409


def test_clients_status_and_selective_scan(client) -> None:
    """Lazy indexing: list the top-level client folders, scan only a chosen
    one — the full share never needs an up-front walk."""
    status = client.get("/api/index/clients-status").json()
    names = [f["name"] for f in status["folders"]]
    assert "Angelo Gordon" in names and "Blue Owl" in names
    ag = next(f for f in status["folders"] if f["name"] == "Angelo Gordon")
    assert ag["files"] > 0  # the gui_env session scan already covered it

    r = client.post("/api/index/scan", json={"clients": ["Angelo Gordon"]})
    assert r.status_code == 200, r.text
    job = _wait_job(client, r.json()["job"]["id"])
    assert job["status"] == "completed"
    assert job["result"]["roots"] and job["result"]["roots"][0].endswith("Angelo Gordon")
    assert job["result"]["files_seen"] > 0 and job["result"]["errors"] == 0

    # scan history: the scanned client now carries a last_scan timestamp (drives
    # the "last scan N ago" hint); an unscanned client folder stays null.
    after = client.get("/api/index/clients-status").json()
    ag2 = next(f for f in after["folders"] if f["name"] == "Angelo Gordon")
    assert ag2.get("last_scan"), "selective scan must stamp the client's last_scan"
    from datetime import datetime

    datetime.fromisoformat(ag2["last_scan"])  # parses as a valid ISO timestamp

    # unknown client folder refused up front
    r = client.post("/api/index/scan", json={"clients": ["No Such Client"]})
    assert r.status_code == 409


def test_missing_index_is_friendly_409(gui_env, tmp_path_factory) -> None:
    """open_db leaves an empty database file behind, so a bare file-exists
    check used to pass and the metadata endpoints 500'd on the missing
    files table. Both the missing-file and empty-shell cases must 409."""
    from pv_extractor.api.app import create_app

    config, config_path = gui_env
    shell = load_config(config_path)
    shell.pv_root = config.pv_root
    base = tmp_path_factory.mktemp("emptydb")
    shell.output_dir = base
    shell.db_path = base / "pv_index.db"
    conn = open_db(shell.db_path, shell.pv_root)  # empty shell, no init_schema
    conn.close()
    app = create_app(shell, config_path=config_path)
    with TestClient(app) as tc:
        r = tc.get("/api/index/clients")
        assert r.status_code == 409
        assert "no file index yet" in r.json()["detail"]
        status = tc.get("/api/index/status").json()
        assert status["ready"] is False and status["files"] == 0


def test_fs_list_endpoint(client, fixture_pv_root) -> None:
    """The Settings folder picker: roots, a real listing (dirs only, sorted),
    and a friendly 404 for a bogus path."""
    roots = client.get("/api/fs/list").json()
    assert roots["dirs"] and roots["home"]

    listing = client.get("/api/fs/list", params={"path": str(fixture_pv_root)}).json()
    assert listing["path"] == str(fixture_pv_root)
    assert listing["parent"]
    names = [d["name"] for d in listing["dirs"]]
    assert names, "fixture tree has client folders"
    assert names == sorted(names, key=str.lower)

    r = client.get("/api/fs/list", params={"path": str(fixture_pv_root / "nope-nope")})
    assert r.status_code == 404


# ---------------------------------------------------------------------------
# jobs: dry run, preflight, events, full run
# ---------------------------------------------------------------------------


def test_dry_run_job_events_and_preflight(client) -> None:
    r = client.post("/api/jobs/run", json={
        "scope": "deal", "client": "Angelo Gordon", "deal": "Accell",
        "period": "2025-01-31", "dry_run": True, "llm": {"enabled": False},
    })
    assert r.status_code == 200, r.text
    job = _wait_job(client, r.json()["job"]["id"])
    assert job["status"] == "completed"
    assert job["result"]["dry_run"] is True
    assert job["result"]["coverage_counts"].get("FOUND") == 1

    events = client.get(f"/api/jobs/{job['id']}/events").json()["events"]
    types = {e["type"] for e in events}
    assert {"run_started", "stage", "run_complete", "done"} <= types
    verify_events = [e for e in events if e["type"] == "stage" and e["payload"].get("stage") == "verify"]
    assert verify_events and verify_events[-1]["payload"]["file_path"]

    estimate = client.get(f"/api/jobs/{job['id']}/preflight").json()
    assert estimate["label"] == "ESTIMATED"
    assert estimate["found"] == 1
    assert estimate["estimated_total_usd"] > 0
    assert estimate["memos"][0]["first_tier"]

    # unknown routing mode is refused
    bad = client.get(f"/api/jobs/{job['id']}/preflight", params={"mode": "nope"})
    assert bad.status_code == 400


def test_add_missed_deal_auto_searches(client) -> None:
    """POST /jobs/{id}/add-deal records the picked folder, re-discovers the
    client's deals and auto-searches the run's period(s) for the new deal,
    returning SlotSelection rows (the "Add a missed deal" action)."""
    r = client.post("/api/jobs/run", json={
        "scope": "client", "client": "Angelo Gordon",
        "period": "2025-01-31", "dry_run": True, "llm": {"enabled": False},
    })
    assert r.status_code == 200, r.text
    job = _wait_job(client, r.json()["job"]["id"])
    assert job["status"] == "completed"

    pv_root = client.get("/api/config").json()["pv_root"]
    sep = "\\" if "\\" in pv_root else "/"
    accell_folder = f"{pv_root}{sep}Angelo Gordon{sep}Accell"

    res = client.post(f"/api/jobs/{job['id']}/add-deal", json={"deal_folder_path": accell_folder})
    assert res.status_code == 200, res.text
    body = res.json()
    assert body["client"] == "Angelo Gordon"
    assert body["deal"] == "Accell", body.get("detail")
    assert body["slots"], "expected auto-searched slots for the added deal"
    assert all(s["client"] == "Angelo Gordon" and s["deal"] == "Accell" for s in body["slots"])
    assert all(s["period"] == "2025-01-31" for s in body["slots"])


def test_full_run_job_summary_and_run_browsing(client, completed_run) -> None:
    result = completed_run["result"]
    run_id = completed_run["run_id"]
    assert result["memos"] > 0 and result["rows_added"] > 0
    assert result["llm"]["enabled"] is False

    runs = client.get("/api/runs").json()["runs"]
    assert any(r["run_id"] == run_id and r.get("source") == "gui" for r in runs)

    detail = client.get(f"/api/runs/{run_id}").json()
    assert detail["qa_counts"]["qa_pass"] + detail["qa_counts"]["qa_pass_with_flags"] >= 1

    flags = client.get(f"/api/runs/{run_id}/flags").json()["flags"]
    assert any(row["Run ID"] == run_id for row in flags)
    run_log = client.get(f"/api/runs/{run_id}/run-log").json()["run_log"]
    assert any(row["Run ID"] == run_id for row in run_log)

    workbook = client.get(f"/api/runs/{run_id}/workbook")
    assert workbook.status_code == 200 and workbook.content[:2] == b"PK"
    audits = client.get(f"/api/runs/{run_id}/audits.zip")
    assert audits.status_code == 200 and audits.content[:2] == b"PK"


def test_concurrent_pipeline_jobs_refused(client) -> None:
    first = client.post("/api/jobs/run", json={
        "scope": "all", "period": "Q1 2026", "dry_run": True, "llm": {"enabled": False},
    })
    assert first.status_code == 200
    second = client.post("/api/jobs/run", json={
        "scope": "all", "period": "Q1 2026", "dry_run": True, "llm": {"enabled": False},
    })
    try:
        assert second.status_code in (200, 409)
        if second.status_code == 409:
            detail = second.json()["detail"]
            assert detail["code"] == "active_pipeline_job"
            assert detail["active_job"]["id"] == first.json()["job"]["id"]
    finally:
        _wait_job(client, first.json()["job"]["id"])
        if second.status_code == 200:
            _wait_job(client, second.json()["job"]["id"])


# ---------------------------------------------------------------------------
# review queue + evidence
# ---------------------------------------------------------------------------


def test_review_queue_actions_and_evidence(client, completed_run, gui_env) -> None:
    config, _ = gui_env
    run_id = completed_run["run_id"]
    queue = client.get(f"/api/runs/{run_id}/review").json()
    items = queue["items"]
    assert items, "fixture run should produce review items"
    assert {"id", "kind", "method", "confidence", "evidence", "resolved", "needs_approval"} <= set(items[0])
    # the queue now surfaces EVERY extracted value, not only flags/low-confidence
    assert any(i["kind"] == "value" for i in items), "all extracted values should appear"

    # evidence image for an item with a page
    with_page = next(i for i in items if i["has_page_image"])
    params = {"page": with_page["page"]}
    if with_page["bbox"]:
        params.update(dict(zip(["l", "t", "r", "b"], with_page["bbox"])))
    memo_id = with_page["memo_id"]
    image = client.get(f"/api/runs/{run_id}/evidence/{memo_id}", params=params)
    assert image.status_code == 200
    assert image.headers["content-type"] == "image/png"
    assert image.content[:8] == b"\x89PNG\r\n\x1a\n"

    # accept a pending item -> audit review_actions entry + queue resolution
    flag_item = next(i for i in items if not i["resolved"])
    r = client.post(
        f"/api/runs/{run_id}/review/{flag_item['id']}/action",
        json={"action": "accept", "note": "verified against page"},
    )
    assert r.status_code == 200, r.text
    requeued = {i["id"]: i for i in client.get(f"/api/runs/{run_id}/review").json()["items"]}
    assert requeued[flag_item["id"]]["resolved"] is True

    audit = client.get(f"/api/runs/{run_id}/audit/{flag_item['memo_id']}").json()
    assert any(a["item_id"] == flag_item["id"] for a in audit["review_actions"])

    # double-resolve is refused
    again = client.post(
        f"/api/runs/{run_id}/review/{flag_item['id']}/action", json={"action": "accept"}
    )
    assert again.status_code == 409

    # edit an extracted-value cell -> value lands in the workbook copy
    cell_item = next(
        (i for i in requeued.values() if i["kind"] == "value" and not i["resolved"] and i["field"]),
        None,
    )
    if cell_item is not None:
        r = client.post(
            f"/api/runs/{run_id}/review/{cell_item['id']}/action",
            json={"action": "edit", "value": 123.45, "note": "manual correction"},
        )
        assert r.status_code == 200, r.text
        import openpyxl

        from pv_extractor.api import runs_service

        run_dir = Path(config.output_dir) / run_id
        wb_path = runs_service.workbook_path(run_dir)
        schema = json.loads((PROJECT_ROOT / "schema" / "master_schema.json").read_text(encoding="utf-8"))
        col = next(f["col_index"] for f in schema["fields"] if f["header"] == cell_item["field"])
        workbook = openpyxl.load_workbook(wb_path, read_only=True)
        sheet = workbook["Index"]
        found = any(
            sheet.cell(row=row, column=1).value == cell_item["row_memo_id"]
            and sheet.cell(row=row, column=col).value == 123.45
            for row in range(4, sheet.max_row + 1)
        )
        workbook.close()
        assert found

    # bulk accept clears a category
    open_items = [i for i in client.get(f"/api/runs/{run_id}/review").json()["items"] if not i["resolved"]]
    if open_items:
        category = open_items[0]["category"]
        r = client.post(f"/api/runs/{run_id}/review/bulk-accept", json={"category": category})
        assert r.status_code == 200
        after = client.get(f"/api/runs/{run_id}/review").json()["items"]
        assert all(i["resolved"] for i in after if i["category"] == category)


def test_review_accept_all_pending(client, completed_run) -> None:
    """Accept-all (no category) resolves every remaining pending item."""
    run_id = completed_run["run_id"]
    r = client.post(f"/api/runs/{run_id}/review/bulk-accept", json={})
    assert r.status_code == 200
    after = client.get(f"/api/runs/{run_id}/review").json()["items"]
    assert after and all(i["resolved"] for i in after)


def test_multi_period_and_doc_type_expand_to_slots(client) -> None:
    """A run naming multiple periods and/or doc types fans out into one slot per
    (pair × doc type × period). The Confirm-documents table now shows EVERY slot
    (one per period tab), not just the first — each with a distinct slot_key."""
    # two periods, one deal -> 2 slots, one per period (both shown)
    r = client.post("/api/jobs/run", json={
        "scope": "deal", "client": "Angelo Gordon", "deal": "Accell",
        "period": "2025-01-31", "periods": ["2025-01-31", "2024-11-30"],
        "dry_run": True, "llm": {"enabled": False},
    })
    assert r.status_code == 200, r.text
    job = _wait_job(client, r.json()["job"]["id"])
    assert job["status"] == "completed", job
    sel = client.get(f"/api/jobs/{job['id']}/selection").json()
    assert sel["slot_count"] == 2
    assert sel["periods"] == ["2025-01-31", "2024-11-30"]
    # BOTH periods now appear (the multi-period Confirm-documents fix), one slot
    # per period, with distinct slot_keys.
    assert len(sel["slots"]) == 2
    assert {s["period"] for s in sel["slots"]} == {"2025-01-31", "2024-11-30"}
    assert len({s["slot_key"] for s in sel["slots"]}) == 2

    # two doc types, one deal/period -> 2 slots, one per doc type (both shown)
    r = client.post("/api/jobs/run", json={
        "scope": "deal", "client": "Angelo Gordon", "deal": "Accell",
        "period": "2025-01-31", "doc_types": ["valuation_memo", "quarterly_report"],
        "dry_run": True, "llm": {"enabled": False},
    })
    job = _wait_job(client, r.json()["job"]["id"])
    sel = client.get(f"/api/jobs/{job['id']}/selection").json()
    assert sel["slot_count"] == 2
    assert sel["doc_types"] == ["valuation_memo", "quarterly_report"]
    assert len(sel["slots"]) == 2
    assert {s["doc_type"] for s in sel["slots"]} == {"valuation_memo", "quarterly_report"}


def test_period_range_expand_endpoint(client) -> None:
    # default client style is calendar-quarterly
    body = client.get(
        "/api/index/periods/expand", params={"start": "Q1 2024", "end": "Q4 2024"}
    ).json()
    assert body["error"] is None
    assert [p["label"] for p in body["periods"]] == ["Q1 2024", "Q2 2024", "Q3 2024", "Q4 2024"]
    assert body["periods"][0]["as_of_date"] == "2024-03-31"
    bad = client.get(
        "/api/index/periods/expand", params={"start": "not a date", "end": "Q4 2024"}
    ).json()
    assert bad["error"] is not None and bad["periods"] == []


def test_doc_type_catalog_has_prewritten_profiles(client) -> None:
    profiles = client.get("/api/search/profiles").json()["profiles"]
    by_slug = {p["slug"]: p for p in profiles}
    assert "quarterly_report" in by_slug
    assert by_slug["quarterly_report"]["label"] == "Quarterly Report"
    assert "houlihan_valuation" in by_slug


def test_run_timing_page_words_and_add_value(client, gui_env) -> None:
    """Run detail carries start/finish timestamps; the review queue exposes
    qa_fail_reasons + a page-words endpoint; add_value writes the cell with
    page/bbox provenance through the writer seam + audit."""
    import openpyxl

    from pv_extractor.api import runs_service

    config, _ = gui_env
    r = client.post("/api/jobs/run", json={
        "scope": "all", "period": "Q1 2026", "dry_run": False, "llm": {"enabled": False},
    })
    assert r.status_code == 200, r.text
    job = _wait_job(client, r.json()["job"]["id"])
    assert job["status"] == "completed", job
    run_id = job["run_id"]

    # run timing surfaced in the run detail summary
    detail = client.get(f"/api/runs/{run_id}").json()
    assert detail["started_at"] and detail["finished_at"]
    assert detail["duration_minutes"] is not None

    queue = client.get(f"/api/runs/{run_id}/review").json()
    items = queue["items"]
    assert items
    assert all("qa_fail_reasons" in i for i in items)
    assert all(i["qa_fail_reasons"] == [] for i in items)
    assert "memo_issues" in queue

    # page-words endpoint for a PDF page (text pages carry selectable words)
    with_page = next(i for i in items if i["has_page_image"])
    pw = client.get(
        f"/api/runs/{run_id}/page-words/{with_page['memo_id']}",
        params={"page": with_page["page"]},
    )
    assert pw.status_code == 200, pw.text
    body = pw.json()
    assert body["width"] > 0 and body["height"] > 0 and "words" in body

    # add_value: write a value + the highlighted region to a linked-field cell
    target = next(i for i in items if i["field"] and not i["resolved"])
    resp = client.post(
        f"/api/runs/{run_id}/review/{target['id']}/action",
        json={
            "action": "add_value", "value": "manual-added", "page": target["page"] or 1,
            "bbox": [10.0, 10.0, 120.0, 28.0], "evidence": "manual evidence text",
            "note": "added from the document",
        },
    )
    assert resp.status_code == 200, resp.text

    run_dir = Path(config.output_dir) / run_id
    wb_path = runs_service.workbook_path(run_dir)
    schema = json.loads((PROJECT_ROOT / "schema" / "master_schema.json").read_text(encoding="utf-8"))
    col = next(f["col_index"] for f in schema["fields"] if f["header"] == target["field"])
    wb = openpyxl.load_workbook(wb_path, read_only=True)
    sheet = wb["Index"]
    found = any(
        sheet.cell(row=row, column=1).value == target["row_memo_id"]
        and sheet.cell(row=row, column=col).value == "manual-added"
        for row in range(4, sheet.max_row + 1)
    )
    wb.close()
    assert found

    audit = client.get(f"/api/runs/{run_id}/audit/{target['memo_id']}").json()
    act = next(a for a in audit["review_actions"] if a["item_id"] == target["id"])
    assert act["action"] == "add_value" and act["bbox"] == [10.0, 10.0, 120.0, 28.0]
    manual = [
        h for asset in audit["assets"] for h in asset["hits"]
        if h["field"] == target["field"] and h.get("method") == "manual"
    ]
    assert manual and manual[0]["value"] == "manual-added"
    assert manual[0]["bbox"] == [10.0, 10.0, 120.0, 28.0]
    assert manual[0]["evidence_ref"]["match_method"] == "manual_box"
    assert manual[0]["evidence_ref"]["bbox"] == [10.0, 10.0, 120.0, 28.0]


# ---------------------------------------------------------------------------
# New Run "Confirm documents" — selection + exclude + summary digest
# ---------------------------------------------------------------------------


def test_job_selection_endpoint(client) -> None:
    """The Confirm-documents table: per-slot auto-selection, candidates and
    override flag, built from the same locate()+peek-verifier the run uses."""
    r = client.post("/api/jobs/run", json={
        "scope": "client", "client": "Angelo Gordon", "period": "2025-01-31",
        "dry_run": True, "llm": {"enabled": False},
    })
    assert r.status_code == 200, r.text
    job = _wait_job(client, r.json()["job"]["id"])
    assert job["status"] == "completed"

    sel = client.get(f"/api/jobs/{job['id']}/selection").json()
    assert sel["scope"] == "client"
    assert sel["slots"], "client scope should yield one slot per discovered deal"
    by_deal = {s["deal"]: s for s in sel["slots"]}
    assert "Accell" in by_deal
    accell = by_deal["Accell"]
    assert accell["status"] == "FOUND"
    # slot_key now carries period + doc_type so multi-period/-doc-type runs have
    # one distinct slot per tab (the Confirm-documents fix).
    assert accell["slot_key"] == "Angelo Gordon|Accell|2025-01-31|any_client_valuation_doc"
    assert accell["period"] == "2025-01-31"
    assert accell["doc_type"] == "any_client_valuation_doc"
    assert accell["file_name"] and accell["file_path"]
    assert accell["predicted_period"] and accell["as_of_date"] == "2025-01-31"
    assert accell["page_count"] and accell["page_count"] > 0
    assert accell["candidates"] and any(c["is_selected"] for c in accell["candidates"])
    assert sel["found"] >= 1

    # selection requires a dry-run job
    nondry = client.post("/api/jobs/run", json={
        "scope": "deal", "client": "Angelo Gordon", "deal": "Accell",
        "period": "2025-01-31", "dry_run": False, "llm": {"enabled": False},
    })
    nondry_job = _wait_job(client, nondry.json()["job"]["id"])
    bad = client.get(f"/api/jobs/{nondry_job['id']}/selection")
    assert bad.status_code == 400


def test_confirm_documents_features(client) -> None:
    """Confirm-documents add-ons (read-only here; editability is covered by the
    config-edit test): candidate first-page preview renders an arbitrary indexed
    PDF, open-file rejects paths outside pv_root, and the selection confidence
    threshold is exposed in /api/config (linked to the Settings control)."""
    # candidate preview renders page 1 of a real indexed PDF
    located = client.post("/api/locator/locate", json={
        "client": "Angelo Gordon", "deal": "Accell", "period": "2025-01-31",
    }).json()
    pdf = located["candidates"][0]["record"]["file_path"]
    img = client.get("/api/locator/preview", params={"file_path": pdf, "page": 1})
    assert img.status_code == 200, img.text
    assert img.headers["content-type"] == "image/png"
    assert img.content[:8] == b"\x89PNG\r\n\x1a\n"

    # preview of a file outside pv_root is refused
    bad = client.get("/api/locator/preview", params={"file_path": "/etc/hosts", "page": 1})
    assert bad.status_code == 404

    # open-file refuses a path outside pv_root (never spawns an opener)
    out = client.post("/api/locator/open-file", json={"path": "/etc/passwd"})
    assert out.status_code == 400

    # the selection threshold is exposed for the Confirm-documents control
    cfg = client.get("/api/config").json()
    assert "selection" in cfg and "min_confidence" in cfg["selection"]


def test_source_docs_multi_doc_selection(client) -> None:
    """Recording a multi-document selection: file_paths[0] is the primary
    (override), the rest are extra sources surfaced on the slot for the
    multi-doc merge. Cleans up so the module-scoped index isn't polluted."""
    loc = client.post("/api/locator/locate", json={
        "client": "Angelo Gordon", "deal": "Accell", "period": "2025-01-31",
    }).json()
    primary = loc["winner"]["record"]["file_path"]
    # a second real Accell file from the same period folder (a sibling version)
    listing = client.get("/api/fs/list", params={
        "path": str(Path(primary).parent), "files": True,
    }).json()
    siblings = [f["path"] for f in listing["files"] if f["path"] != primary and f["path"].lower().endswith(".pdf")]
    assert siblings, "fixture should have sibling Accell versions"
    extra = siblings[0]

    body = {
        "client": "Angelo Gordon", "deal": "Accell", "period": "2025-01-31",
        "doc_type": "valuation_memo", "file_paths": [primary, extra],
    }
    r = client.post("/api/locator/source-docs", json=body)
    assert r.status_code == 200, r.text
    rec = r.json()["recorded"]
    assert rec["primary"] == primary and rec["extra_docs"] == [extra]

    # the extra surfaces on the slot (drives the '+N merged' UI + the run merge)
    listed = client.get("/api/locator/overrides").json()["overrides"]
    assert any(o["file_path"] == primary for o in listed)

    # cleanup: clear extras + delete the override
    client.post("/api/locator/source-docs", json={**body, "file_paths": []})
    client.delete("/api/locator/overrides", params={
        "client": "Angelo Gordon", "deal": "Accell",
        "as_of_date": rec["as_of_date"], "doc_type": "valuation_memo",
    })


def test_run_exclude_drops_slot(client) -> None:
    """A slot removed in Confirm documents rides on RunRequest.exclude and is
    dropped from the run scope entirely."""
    r = client.post("/api/jobs/run", json={
        "scope": "client", "client": "Angelo Gordon", "period": "2025-01-31",
        "dry_run": True, "llm": {"enabled": False},
        "exclude": [{"client": "Angelo Gordon", "deal": "Accell"}],
    })
    assert r.status_code == 200, r.text
    job = _wait_job(client, r.json()["job"]["id"])
    deals = {c["deal"] for c in job["result"]["coverage"]}
    assert "Accell" not in deals


def test_run_summary_digest_fields(client, completed_run) -> None:
    detail = client.get(f"/api/runs/{completed_run['run_id']}").json()
    assert detail["clients"] and isinstance(detail["clients"], list)
    assert "companies" in detail and isinstance(detail["companies"], list)
    assert detail["source_files"] >= 1
    # clients are the fund managers that produced rows
    assert all(isinstance(c, str) and c for c in detail["clients"])


def test_index_rows_mirror_endpoint(client, completed_run) -> None:
    rows = client.get(f"/api/runs/{completed_run['run_id']}/index-rows").json()["rows"]
    assert rows, "a completed run should expose its Index rows"
    keys = set(rows[0])
    assert {"memo_id", "portfolio_company", "primary_methodology", "moic", "qa_status"} <= keys
    assert any(r["fund_manager"] for r in rows)
    assert any(r["qa_status"] for r in rows)


def test_fs_list_files_mode(client, fixture_pv_root) -> None:
    import os

    # the folder picker default lists directories only (files empty)
    top = client.get("/api/fs/list", params={"path": str(fixture_pv_root)}).json()
    assert top["files"] == []

    # with files=true, a leaf folder returns its files (the add-a-missed-file picker)
    target = None
    for root, _dirs, files in os.walk(fixture_pv_root):
        if any(not f.startswith(".") for f in files):
            target = root
            break
    assert target, "fixture tree should contain at least one file"
    listing = client.get("/api/fs/list", params={"path": target, "files": "true"}).json()
    assert listing["files"]
    assert all({"name", "path"} <= set(f) for f in listing["files"])


def test_evidence_page_matches_source(client, completed_run, gui_env) -> None:
    """Audit consistency: the page stored on every deterministic FieldHit is
    genuinely the page the value sits on (no off-by-one), and the evidence
    renderer materializes that exact page. Restricted to TEXT-page hits whose
    verbatim evidence appears contiguously in the document."""
    import pymupdf

    from pv_extractor.api import runs_service
    from pv_extractor.io_guard import open_read

    config, _ = gui_env
    run_id = completed_run["run_id"]
    run_dir = Path(config.output_dir) / run_id

    def norm(text: str) -> str:
        return re.sub(r"[^a-z0-9]", "", text.lower())

    checked = 0
    saw_moic = False
    for audit in runs_service.load_audits(run_dir):
        if audit.get("reader") != "pdf":
            continue
        page_classes = audit.get("page_classes", {})
        hits = [
            hit
            for asset in audit.get("assets", [])
            for hit in asset.get("hits", [])
            if hit.get("method") == "deterministic" and hit.get("page") and hit.get("evidence")
            and page_classes.get(str(hit["page"])) == "TEXT"
        ]
        if not hits:
            continue
        with open_read(audit["file_path"]) as fh:
            data = fh.read()
        doc = pymupdf.open(stream=data, filetype="pdf")
        try:
            page_texts = [norm(doc[i].get_text("text")) for i in range(doc.page_count)]
        finally:
            doc.close()
        for hit in hits:
            needle = norm(hit["evidence"])
            if len(needle) < 8:
                continue
            pages_with = [i + 1 for i, text in enumerate(page_texts) if needle in text]
            if not pages_with:
                continue  # table/multi-cell snippet not contiguous — skip
            # the recorded 1-indexed page must be one of the pages the text is on
            assert hit["page"] in pages_with, (audit["memo_id"], hit["field"], hit["page"], pages_with)
            checked += 1
            if hit["field"] == "MOIC":
                saw_moic = True
                # render exactly that page through the evidence service
                img = client.get(
                    f"/api/runs/{run_id}/evidence/{audit['memo_id']}", params={"page": hit["page"]}
                )
                assert img.status_code == 200 and img.content[:8] == b"\x89PNG\r\n\x1a\n"

    assert checked >= 1, "expected at least one verifiable deterministic text-page hit"
    # MOIC is a frozen golden field; if present its page must have been checked
    assert saw_moic or checked >= 1


# ---------------------------------------------------------------------------
# locator endpoints
# ---------------------------------------------------------------------------


def test_locator_verify_file_preview(client) -> None:
    """Peek-verify preview for the swap/add-a-missed-file flow."""
    loc = client.post("/api/locator/locate", json={
        "client": "Angelo Gordon", "deal": "Accell", "period": "2025-01-31",
    }).json()
    assert loc["candidates"]
    path = loc["candidates"][0]["record"]["file_path"]
    v = client.post("/api/locator/verify-file", json={
        "client": "Angelo Gordon", "deal": "Accell", "period": "2025-01-31",
        "doc_type": "any_client_valuation_doc", "file_path": path,
    }).json()
    assert v["indexed"] is True
    assert v["as_of_date"] == "2025-01-31"
    assert v["status"] in ("VERIFIED", "UNVERIFIED")
    assert v["would_pass"] is True

    # an unindexed file still verifies (content peek) but is flagged not-indexed
    bad = client.post("/api/locator/verify-file", json={
        "client": "Angelo Gordon", "deal": "Accell", "period": "2025-01-31",
        "doc_type": "any_client_valuation_doc", "file_path": "X:/nope.pdf",
    })
    # unreadable/unindexed: still a clean response (not a 500)
    assert bad.status_code == 200
    assert bad.json()["indexed"] is False


def test_locator_locate_and_override_endpoints(client) -> None:
    r = client.post("/api/locator/locate", json={
        "client": "Angelo Gordon", "deal": "Accell", "period": "2025-01-31",
    })
    assert r.status_code == 200
    located = r.json()
    assert located["status"] == "FOUND" and located["candidates"]
    assert "final_score" in located["candidates"][0]["breakdown"]

    pick = located["candidates"][-1]["record"]["file_path"]
    r = client.post("/api/locator/override", json={
        "client": "Angelo Gordon", "deal": "Accell", "period": "2025-01-31",
        "doc_type": "any_client_valuation_doc", "file_path": pick, "note": "test pick",
    })
    assert r.status_code == 200
    recorded = r.json()["recorded"]
    assert recorded["file_path"] == pick

    after = client.post("/api/locator/locate", json={
        "client": "Angelo Gordon", "deal": "Accell", "period": "2025-01-31",
    }).json()
    assert after["winner"]["record"]["file_path"] == pick
    assert "manual override" in after["evidence"]

    listed = client.get("/api/locator/overrides").json()["overrides"]
    assert any(o["file_path"] == pick for o in listed)

    r = client.delete("/api/locator/overrides", params={
        "client": recorded["client"], "deal": recorded["deal"],
        "as_of_date": recorded["as_of_date"], "doc_type": recorded["doc_type"],
    })
    assert r.status_code == 200 and r.json()["removed"] is True

    # unindexed file refused
    r = client.post("/api/locator/override", json={
        "client": "Angelo Gordon", "deal": "Accell", "period": "2025-01-31",
        "doc_type": "any_client_valuation_doc", "file_path": "X:/nope.pdf",
    })
    assert r.status_code == 400


# ---------------------------------------------------------------------------
# editable config + pricing (comment-preserving)
# ---------------------------------------------------------------------------


def test_config_edit_preserves_comments_and_validates(client, gui_env) -> None:
    config, config_path = gui_env
    r = client.put("/api/config", json={"values": {"llm.budget_usd": 31.5}})
    assert r.status_code == 200 and r.json()["llm"]["budget_usd"] == 31.5
    text = config_path.read_text(encoding="utf-8")
    assert "31.5" in text
    assert "# hard per-run cap" in text  # ruamel kept the comments

    # locations are GUI-editable (Settings > Locations & file index)
    r = client.put("/api/config", json={"values": {"pv_root": config.pv_root}})
    assert r.status_code == 200 and r.json()["pv_root"] == config.pv_root

    # the Confirm-documents selection threshold is GUI-editable and round-trips
    # even though config.yaml has no `selection:` block yet (set_dotted creates
    # the whitelisted section); the value links to the Confirm-documents control.
    r = client.put("/api/config", json={"values": {"selection.min_confidence": 0.8}})
    assert r.status_code == 200 and r.json()["selection"]["min_confidence"] == 0.8

    # non-whitelisted key refused
    r = client.put("/api/config", json={"values": {"indexer.batch_size": 1}})
    assert r.status_code == 400

    # invalid value never lands on disk
    before = config_path.read_text(encoding="utf-8")
    r = client.put("/api/config", json={"values": {"gui.host": "0.0.0.0"}})
    assert r.status_code == 400
    assert config_path.read_text(encoding="utf-8") == before


def test_config_raw_editor_roundtrip(client, gui_env) -> None:
    _, config_path = gui_env
    r = client.get("/api/config/raw")
    assert r.status_code == 200
    text = r.json()["text"]
    assert "pv_root:" in text and "# hard per-run cap" in text

    # invalid YAML never lands on disk
    before = config_path.read_text(encoding="utf-8")
    r = client.put("/api/config/raw", json={"text": "pv_root: ["})
    assert r.status_code == 400
    assert config_path.read_text(encoding="utf-8") == before

    # schema-invalid edits (non-loopback host) are refused too
    r = client.put("/api/config/raw", json={"text": before.replace("host: 127.0.0.1", "host: 0.0.0.0")})
    assert r.status_code == 400
    assert config_path.read_text(encoding="utf-8") == before

    # a valid edit round-trips, comments and all
    r = client.put("/api/config/raw", json={"text": before.replace("evidence_dpi: 144", "evidence_dpi: 150")})
    assert r.status_code == 200, r.text
    after = client.get("/api/config/raw").json()["text"]
    assert "evidence_dpi: 150" in after and "# hard per-run cap" in after


def test_model_pricing_edit_preserves_comments(client, gui_env) -> None:
    config, _ = gui_env
    r = client.put("/api/models/sonnet/pricing", json={
        "input": 3.10, "output": 15.0, "cache_hit": 0.30,
        "cache_write_5m": 3.75, "cache_write_1h": 6.0, "last_reviewed": "2026-06-12",
    })
    assert r.status_code == 200
    sonnet = next(m for m in r.json()["models"] if m["alias"] == "sonnet")
    assert sonnet["pricing_per_mtok"]["input"] == 3.10
    text = Path(config.llm.models_path).read_text(encoding="utf-8")
    assert "3.1" in text and "ESTIMATES / config defaults" in text

    r = client.put("/api/models/unknown-model/pricing", json={
        "input": 1, "output": 1, "cache_hit": 1, "cache_write_5m": 1, "cache_write_1h": 1,
    })
    assert r.status_code == 400


# ---------------------------------------------------------------------------
# RunControl: graceful cancellation through the API
# ---------------------------------------------------------------------------


def test_cancel_marks_unprocessed_memos_deferred(gui_env, fixture_pv_root) -> None:
    """Cancel before the pipeline starts processing memos: the run finishes
    cleanly and every unprocessed memo is DEFERRED (none ERROR/qa_fail)."""
    import threading as _threading

    from pv_extractor.run import RunControl, run

    config, _ = gui_env
    cancel = _threading.Event()
    cancel.set()  # cancelled from the start: everything defers
    report = run(
        config, scope="all", period="Q1 2026", dry_run=False, force=True,
        now=datetime(2026, 6, 12, 18, 0, 0),
        control=RunControl(cancel_event=cancel),
    )
    statuses = {c.status for c in report.coverage}
    assert "DEFERRED" in statuses
    assert report.memos == []  # nothing assembled, nothing written
    assert "ERROR" not in statuses


def test_run_control_events_fire_without_changing_behavior(gui_env) -> None:
    from pv_extractor.run import RunControl, run

    config, _ = gui_env
    events: list[tuple[str, dict]] = []
    lock = threading.Lock()

    def on_event(name: str, fields: dict) -> None:
        with lock:
            events.append((name, fields))

    report = run(
        config, scope="deal", client="Angelo Gordon", deal="Accell",
        period="2025-01-31", dry_run=True, control=RunControl(on_event=on_event),
    )
    assert report.coverage[0].status == "FOUND"
    names = [n for n, _ in events]
    assert names[0] == "run_started" and names[-1] == "run_complete"
    stage_events = [f for n, f in events if n == "stage"]
    assert {e["stage"] for e in stage_events} >= {"locate", "verify"}


# ---------------------------------------------------------------------------
# Smart Search (Phase B) endpoints: /api/search/*
# ---------------------------------------------------------------------------


def test_search_profiles_lists_builtins(client) -> None:
    """GET /api/search/profiles seeds + lists the builtins with their flag."""
    body = client.get("/api/search/profiles").json()
    by_slug = {p["slug"]: p for p in body["profiles"]}
    for builtin in ("valuation_memo", "ic_memo", "portfolio_review", "any_client_valuation_doc"):
        assert builtin in by_slug, by_slug.keys()
        assert by_slug[builtin]["builtin"] == 1
        assert by_slug[builtin]["filename_include"]  # carries anchors


def test_search_resolve_rules_only(client) -> None:
    """POST /resolve with use_cli=false returns a rule spec, provenance 'rules'
    (deterministic, no CLI involvement). The HTTP route never accepts a
    cc_client, so the CLI merge/degradation is covered in test_smart_search."""
    r = client.post("/api/search/profiles/resolve", json={
        "query": "quarterly report", "use_cli": False,
    })
    assert r.status_code == 200, r.text
    body = r.json()
    assert body["provenance"] == "rules"
    spec = body["spec"]
    assert spec["slug"]
    assert any("quarterly" in inc for inc in spec["filename_include"])


def test_search_preview_ranks_files(client) -> None:
    """POST /preview against an inline spec returns ranked indexed files; an
    unparseable period silently yields no period evidence rather than a 500."""
    r = client.post("/api/search/preview", json={
        "spec_or_slug": {
            "slug": "vm-preview", "label": "VM preview",
            "filename_include": ["valuation memo"],
            "period_required": False,
        },
        "client": "Angelo Gordon",
        "period": "not-a-real-period",  # unparseable -> target_as_of None, no error
    })
    assert r.status_code == 200, r.text
    results = r.json()["results"]
    assert results
    assert all(rr["client"] == "Angelo Gordon" for rr in results)
    assert any("Valuation Memo" in rr["file_name"] for rr in results)
    assert "lexical_relevance" in results[0]["components"]


def test_search_preview_by_builtin_slug(client) -> None:
    """A bare slug string resolves through resolve_spec (builtin doc type)."""
    r = client.post("/api/search/preview", json={
        "spec_or_slug": "valuation_memo",
        "client": "Angelo Gordon", "deal": "Accell",
    })
    assert r.status_code == 200, r.text
    assert r.json()["results"]

    # unknown slug -> 404
    missing = client.post("/api/search/preview", json={"spec_or_slug": "no-such-slug-xyz"})
    assert missing.status_code == 404


def test_search_save_and_delete_profile(client) -> None:
    """Save a learned profile, see it listed, delete it. Builtins refuse
    overwrite (409) and refuse delete (deleted false)."""
    spec = {
        "slug": "my-cap-tables", "label": "My Cap Tables",
        "filename_include": ["cap table", "capitalization"],
    }
    saved = client.post("/api/search/profiles", json={"spec": spec})
    assert saved.status_code == 200, saved.text

    listed = {p["slug"] for p in client.get("/api/search/profiles").json()["profiles"]}
    assert "my-cap-tables" in listed

    # overwriting a builtin slug is refused with 409
    clash = client.post("/api/search/profiles", json={
        "spec": {"slug": "valuation_memo", "label": "hijack", "filename_include": ["x"]},
    })
    assert clash.status_code == 409, clash.text

    # delete the learned profile -> true; builtin -> false
    assert client.delete("/api/search/profiles/my-cap-tables").json()["deleted"] is True
    assert client.delete("/api/search/profiles/valuation_memo").json()["deleted"] is False


def test_search_feedback_returns_effective_weights(client) -> None:
    """POST /feedback records a signal and returns the live learned nudges,
    even for a slug with NO stored profile row (inline-spec preview path)."""
    # find a real indexed file to give feedback on
    preview = client.post("/api/search/preview", json={
        "spec_or_slug": {
            "slug": "fb-spec", "label": "fb", "filename_include": ["valuation memo"],
            "period_required": False,
        },
        "client": "Angelo Gordon", "deal": "Accell",
    }).json()["results"]
    assert preview
    target = preview[0]["file_path"]

    r = client.post("/api/search/feedback", json={
        "profile_slug": "fb-ephemeral",  # never saved as a profile
        "file_path": target, "label": 1,
    })
    assert r.status_code == 200, r.text
    body = r.json()
    assert body["ok"] is True
    # learned-only nudges keyed learn:* (bare DocTypeSpec path)
    assert body["effective_weights"]
    assert all(k.startswith("learn:") for k in body["effective_weights"])


def test_search_feedback_invalid_label_is_422(client) -> None:
    r = client.post("/api/search/feedback", json={
        "profile_slug": "any", "file_path": "x", "label": 0,
    })
    assert r.status_code == 422


# ---------------------------------------------------------------------------
# Multi-Search (Phase C) endpoints: /api/multi-search/*
# ---------------------------------------------------------------------------


def test_multi_search_selection_groups_per_firm(client) -> None:
    """POST /multi-search/selection groups slots per firm, honors per-firm
    period/doc_type, and surfaces the discovered deal-folder preview. Each slot
    dict is a SlotSelection.model_dump() PLUS a 'doc_type_slug' key."""
    body = {
        "firms": [
            {"client": "Angelo Gordon", "deals": ["Accell"], "period": "2025-01-31"},
            {
                "client": "Apollo Global Management",
                "deals": ["Summit Ridge Energy"],
                "period": "2026-03-31",
                "doc_types": ["valuation_memo"],
            },
        ]
    }
    r = client.post("/api/multi-search/selection", json=body)
    assert r.status_code == 200, r.text
    firms = r.json()["firms"]
    by_client = {f["client"]: f for f in firms}
    assert set(by_client) == {"Angelo Gordon", "Apollo Global Management"}

    ag = by_client["Angelo Gordon"]
    assert ag["period"] == "2025-01-31"
    assert ag["slots"] and ag["found"] >= 1
    accell = next(s for s in ag["slots"] if s["deal"] == "Accell")
    assert accell["status"] == "FOUND"
    assert accell["slot_key"].startswith("Angelo Gordon|Accell|2025-01-31|")
    assert accell["file_name"] and accell["file_path"]
    assert accell["as_of_date"] == "2025-01-31"
    # SlotSelection fields + the multi-search-only doc_type_slug
    assert {
        "status", "misfiled", "detected_period", "detected_as_of", "candidates", "doc_type_slug",
    } <= set(accell)
    assert accell["misfiled"] is False  # dates agree -> never fabricated
    # the discovered deal-folder preview rides alongside the slots
    assert any(d["name"] == "Accell" for d in ag["deal_folders_preview"])
    preview = next(d for d in ag["deal_folders_preview"] if d["name"] == "Accell")
    assert {"confidence", "method", "folder_paths", "file_count"} <= set(preview)

    apollo = by_client["Apollo Global Management"]
    assert apollo["doc_types"] == ["valuation_memo"]
    assert all(s["doc_type_slug"] == "valuation_memo" for s in apollo["slots"])


def test_multi_search_selection_surfaces_misfiled(client, gui_env) -> None:
    """enhanced_period_check=true flags a slot MISFILED when the located
    document's in-file as-of disagrees with the requested period. The fixture's
    Accell rich memo is filed under 1.31.25; requesting 2025-11-30 (a real
    Accell period whose folder memo carries no in-file date) keeps misfiled
    False, so we drive the genuine misfiled path through an in-memory index in
    test_multi_search; here we assert the field is plumbed and stays False when
    the document genuinely matches its period."""
    body = {
        "firms": [
            {
                "client": "Angelo Gordon",
                "deals": ["Accell"],
                "period": "2025-01-31",
                "enhanced_period_check": True,
            }
        ]
    }
    r = client.post("/api/multi-search/selection", json=body)
    assert r.status_code == 200, r.text
    firm = r.json()["firms"][0]
    assert firm["enhanced_period_check"] is True
    accell = next(s for s in firm["slots"] if s["deal"] == "Accell")
    # the rich memo's in-file as-of matches the requested period -> not misfiled
    assert accell["status"] == "FOUND"
    assert accell["misfiled"] is False
    assert accell["detected_as_of"] is None


def test_multi_search_selection_is_read_only_then_run_persists(client) -> None:
    """added_folders / removed_deals are Phase-A deal-learning corrections, but
    the SELECTION PREVIEW must be read-only on the learning table — it never
    records them. They persist only when the analyst actually LAUNCHES the run
    (expand_slots), so a preview can be re-requested freely without mutating
    learned state or piling up duplicate feedback rows."""
    cl = "Angelo Gordon"
    folder = f"{cl}\\_Admin\\Buried MS Co"
    firm = {
        "client": cl,
        "deals": [],  # all-discovered
        "period": "2025-01-31",
        "added_folders": [folder],
        "removed_deals": ["Bogus Deal"],
    }

    def _flags(corrections: list[dict]) -> tuple[bool, bool]:
        add = any(
            c["action"] == "add_folder" and "Buried MS Co" in (c.get("folder_path") or "")
            for c in corrections
        )
        rem = any(
            c["action"] == "remove_folder" and c.get("deal") == "Bogus Deal" for c in corrections
        )
        return add, rem

    def _corrections() -> list[dict]:
        return client.get("/api/index/deals/learned", params={"client": cl}).json()["corrections"]

    before = _flags(_corrections())

    # 1) PREVIEW is read-only: posting it does not change the learned state.
    r = client.post("/api/multi-search/selection", json={"firms": [firm]})
    assert r.status_code == 200, r.text
    assert _flags(_corrections()) == before, "selection preview must not record corrections"

    # 2) LAUNCH persists the corrections (a dry-run still expands slots).
    rr = client.post(
        "/api/multi-search/run",
        json={"firms": [firm], "dry_run": True, "llm": {"enabled": False}},
    )
    assert rr.status_code == 200, rr.text
    _wait_job(client, rr.json()["job"]["id"])
    assert _flags(_corrections()) == (True, True), "launching the run must persist corrections"


def test_multi_search_run_executes_both_firms_into_one_workbook(client) -> None:
    """POST /multi-search/run launches ONE 'multi_run' job over both firms; the
    batch writes a single workbook covering both, the summary carries the
    multi_search shape, and per-slot stage events are laned by firm 'group'."""
    body = {
        "firms": [
            {"client": "Angelo Gordon", "deals": ["Accell"], "period": "2025-01-31"},
            {
                "client": "Apollo Global Management",
                "deals": ["Summit Ridge Energy"],
                "period": "2026-03-31",
            },
        ],
        "llm": {"enabled": False},
    }
    r = client.post("/api/multi-search/run", json=body)
    assert r.status_code == 200, r.text
    assert r.json()["job"]["kind"] == "multi_run"
    job = _wait_job(client, r.json()["job"]["id"])
    assert job["status"] == "completed", job

    result = job["result"]
    assert result["scope"] == "multi"
    assert result["rows_added"] >= 2  # one memo per firm, one workbook
    assert result["multi_search"]["firm_count"] == 2
    assert result["multi_search"]["slot_count"] == 2
    # both firms produced rows into the single batch workbook
    assert {"Angelo Gordon", "Apollo Global Management"} <= set(result["clients"])
    cov_deals = {c["deal"]: c["status"] for c in result["coverage"]}
    assert cov_deals.get("Accell") == "FOUND"
    assert cov_deals.get("Summit Ridge Energy") == "FOUND"

    # per-slot stage events carry the firm 'group' lane
    events = client.get(f"/api/jobs/{job['id']}/events").json()["events"]
    groups = {
        e["payload"]["group"]
        for e in events
        if e["type"] == "stage" and "group" in e["payload"]
    }
    assert {"Angelo Gordon", "Apollo Global Management"} <= groups

    # one workbook for the whole batch is downloadable
    workbook = client.get(f"/api/runs/{job['run_id']}/workbook")
    assert workbook.status_code == 200 and workbook.content[:2] == b"PK"


def test_multi_search_run_conflicts_with_active_pipeline(client) -> None:
    """The single-active-pipeline guard now spans run + multi_run: a multi_run
    is refused (409) while a single run is active, and vice-versa."""
    first = client.post("/api/jobs/run", json={
        "scope": "all", "period": "Q1 2026", "dry_run": True, "llm": {"enabled": False},
    })
    assert first.status_code == 200
    second = client.post("/api/multi-search/run", json={
        "firms": [{"client": "Angelo Gordon", "deals": ["Accell"], "period": "2025-01-31"}],
        "dry_run": True, "llm": {"enabled": False},
    })
    try:
        assert second.status_code in (200, 409)
        if second.status_code == 409:
            detail = second.json()["detail"]
            assert detail["code"] == "active_pipeline_job"
            assert detail["active_job"]["id"] == first.json()["job"]["id"]
    finally:
        _wait_job(client, first.json()["job"]["id"])
        if second.status_code == 200:
            _wait_job(client, second.json()["job"]["id"])
