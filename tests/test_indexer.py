"""Tests for the D3 file-index store: xlsx ingest with full re-derivation,
filesystem scan with incremental refresh, FTS5 prefilter, and batching."""

from __future__ import annotations

import os
import sqlite3
import sys
import time
from datetime import date, datetime
from pathlib import Path
from types import SimpleNamespace

import pytest
from openpyxl import load_workbook

from pv_extractor.config import Config, IndexerConfig
from pv_extractor.indexer.db import (
    as_of_dates_for_deal,
    count_files,
    deals_for_client,
    distinct_clients,
    file_meta_under,
    fts_candidates,
    init_schema,
    insert_records,
    open_db,
    record_from_row,
    scan_errors_under,
)
from pv_extractor.indexer.derive import derive_record
from pv_extractor.indexer.ingest_xlsx import ingest_xlsx
from pv_extractor.indexer.scan_tree import scan_tree, stat_is_cloud_placeholder
from pv_extractor.io_guard import ReadOnlyViolation
from pv_extractor.models import SourceClass
from pv_extractor.normalize import split_path_segments

PROD_PV_ROOT = "\\\\hlhz\\dfs\\nyfva\\PV"
ANGELES_NDA_PATH = (
    "\\\\hlhz\\dfs\\nyfva\\PV\\Angeles Investments\\Global Admin"
    "\\HL - Angeles NDA 28Sep22 v1 to HL (HL 9-30-22) (002).doc"
)
ACCELL_REPORT_PATH = (
    "\\\\hlhz\\dfs\\nyfva\\PV\\Angelo Gordon\\Accell\\1.31.2025\\Analysis"
    "\\Accell 1.31.2025 Report v1.pdf"
)


def _open_fresh_db(db_path: Path, config: Config) -> sqlite3.Connection:
    conn = open_db(db_path, config.pv_root)
    init_schema(conn)
    return conn


def _row_for(conn: sqlite3.Connection, file_path: str) -> sqlite3.Row:
    row = conn.execute("SELECT * FROM files WHERE file_path = ?", (file_path,)).fetchone()
    assert row is not None, f"row not found: {file_path}"
    return row


@pytest.fixture(scope="module")
def prod_config(tmp_path_factory: pytest.TempPathFactory) -> Config:
    out = tmp_path_factory.mktemp("ingest_out")
    return Config(pv_root=PROD_PV_ROOT, output_dir=out, db_path=out / "pv_index.db")


@pytest.fixture(scope="module")
def ingested(prod_config: Config, strip_xlsx_path: Path):
    """The reference export (999 rows) ingested once for this module."""
    conn = _open_fresh_db(prod_config.db_path, prod_config)
    n = ingest_xlsx(conn, strip_xlsx_path, prod_config)
    yield conn, n
    conn.close()


# --------------------------------------------------------------------------
# Ingest + re-derive (D6 requirement)
# --------------------------------------------------------------------------


def test_ingest_count(ingested) -> None:
    conn, n = ingested
    assert n == 999
    assert count_files(conn) == 999
    assert scan_errors_under(conn, PROD_PV_ROOT) == []


def test_corrupt_parent_folder_is_fixed(ingested, strip_xlsx_path: Path) -> None:
    conn, _ = ingested
    workbook = load_workbook(strip_xlsx_path, read_only=True, data_only=True)
    rows = workbook.worksheets[0].iter_rows(values_only=True)
    header = list(next(rows))
    col = {name: i for i, name in enumerate(header)}
    source_corrupt = sum(1 for r in rows if r[col["parent_folder"]] == "#NAME?")
    workbook.close()
    assert source_corrupt == 349  # the source really is corrupt

    bad = conn.execute("SELECT COUNT(*) FROM files WHERE parent_folder = '#NAME?'").fetchone()[0]
    assert bad == 0
    for row in conn.execute("SELECT file_path, folder_path, parent_folder FROM files"):
        assert row["parent_folder"] == split_path_segments(row["folder_path"])[-1]
        assert row["file_path"].startswith(row["folder_path"])


def test_rederived_columns_match_export_convention(ingested, strip_xlsx_path: Path) -> None:
    """Every re-derived mirror column equals the export's own value on every
    clean row (and on every row for the columns the export got right)."""
    conn, _ = ingested
    db_rows = {row["file_path"]: row for row in conn.execute("SELECT * FROM files")}
    workbook = load_workbook(strip_xlsx_path, read_only=True, data_only=True)
    rows = workbook.worksheets[0].iter_rows(values_only=True)
    header = list(next(rows))
    col = {name: i for i, name in enumerate(header)}
    checked = 0
    for r in rows:
        if not r[col["file_path"]]:
            continue
        db = db_rows[r[col["file_path"]]]
        for name in (
            "file_name",
            "folder_path",
            "extension",
            "depth_from_pv_root",
            "normalized_file_name",
            "normalized_folder_path",
            "normalized_full_path",
        ):
            assert db[name] == r[col[name]], (name, r[col["file_path"]])
        for name in ("contains_q4_2025_signal", "contains_q1_2026_signal"):
            assert bool(db[name]) == bool(r[col[name]]), (name, r[col["file_path"]])
        checked += 1
    workbook.close()
    assert checked == 999


def test_spot_check_angeles_nda_row(ingested) -> None:
    conn, _ = ingested
    row = _row_for(conn, ANGELES_NDA_PATH)
    assert row["normalized_file_name"] == "hl angeles nda 28sep22 v1 to hl hl 9 30 22 002 doc"
    assert (
        row["normalized_folder_path"] == "hlhz dfs nyfva pv angeles investments global admin"
    )
    assert row["depth_from_pv_root"] == 2
    assert row["parent_folder"] == "Global Admin"
    assert row["extension"] == ".doc"
    assert row["client"] == "Angeles Investments"
    record = record_from_row(row)
    assert record.version_signal is not None
    assert (record.version_signal.rank, record.version_signal.version_number) == (2, 1)


def test_spot_check_accell_row(ingested) -> None:
    conn, _ = ingested
    record = record_from_row(_row_for(conn, ACCELL_REPORT_PATH))
    assert record.client == "Angelo Gordon"
    assert record.deal == "Accell"
    assert record.date_folder == "1.31.2025"
    assert record.as_of_date == date(2025, 1, 31)
    assert record.source_class is SourceClass.analysis
    assert record.depth_from_pv_root == 4
    assert not record.is_archive
    assert record.modified_time == datetime(2025, 1, 28, 14, 12, 51)


def test_distinct_clients_deals_and_dates(ingested) -> None:
    conn, _ = ingested
    assert distinct_clients(conn) == ["Angeles Investments", "Angelo Gordon"]
    deals = deals_for_client(conn, "Angelo Gordon")
    assert "Accell" in deals
    assert deals == sorted(deals)
    dates = as_of_dates_for_deal(conn, "Angelo Gordon", "Accell")
    assert date(2025, 1, 31) in dates
    assert dates == sorted(dates)
    assert as_of_dates_for_deal(conn, "Angelo Gordon", "no-such-deal") == []


def test_fts_candidates_accell_only(ingested) -> None:
    conn, _ = ingested
    records = fts_candidates(conn, '"angelo gordon" AND "accell"', 50)
    assert 0 < len(records) <= 50
    for record in records:
        assert record.client == "Angelo Gordon"
        assert record.deal == "Accell"
        assert record.row_id is not None


def test_record_round_trip(ingested, prod_config: Config) -> None:
    """record_from_row reproduces derive_record output exactly, including
    as_of_date and the VersionSignal."""
    conn, _ = ingested
    stored = record_from_row(_row_for(conn, ACCELL_REPORT_PATH))
    fresh = derive_record(
        ACCELL_REPORT_PATH,
        size_bytes=stored.size_bytes,
        modified_time=stored.modified_time,
        config=prod_config,
    )
    assert stored == fresh.model_copy(update={"row_id": stored.row_id})
    assert stored.as_of_date == date(2025, 1, 31)
    assert stored.version_signal is not None
    assert stored.version_signal.raw == "v1"


def test_batch_size_seven_gives_identical_results(
    ingested, strip_xlsx_path: Path, tmp_path: Path
) -> None:
    conn_default, _ = ingested
    config = Config(pv_root=PROD_PV_ROOT, indexer=IndexerConfig(batch_size=7))
    conn = _open_fresh_db(tmp_path / "batch7.db", config)
    try:
        assert ingest_xlsx(conn, strip_xlsx_path, config) == 999

        def dump(c: sqlite3.Connection) -> list[dict]:
            out = []
            for row in c.execute("SELECT * FROM files ORDER BY file_path"):
                d = dict(row)
                d.pop("id")
                out.append(d)
            return out

        assert dump(conn) == dump(conn_default)
    finally:
        conn.close()


def test_ingest_limit(strip_xlsx_path: Path, tmp_path: Path) -> None:
    config = Config(pv_root=PROD_PV_ROOT)
    conn = _open_fresh_db(tmp_path / "limited.db", config)
    try:
        assert ingest_xlsx(conn, strip_xlsx_path, config, limit=10) == 10
        assert count_files(conn) == 10
    finally:
        conn.close()


def test_reingest_keeps_ids_stable(strip_xlsx_path: Path, tmp_path: Path) -> None:
    config = Config(pv_root=PROD_PV_ROOT)
    conn = _open_fresh_db(tmp_path / "stable.db", config)
    try:
        ingest_xlsx(conn, strip_xlsx_path, config, limit=20)
        before = dict(conn.execute("SELECT file_path, id FROM files"))
        ingest_xlsx(conn, strip_xlsx_path, config, limit=20)
        after = dict(conn.execute("SELECT file_path, id FROM files"))
        assert count_files(conn) == 20
        assert before == after
    finally:
        conn.close()


def test_open_db_refuses_share(tmp_path: Path) -> None:
    with pytest.raises(ReadOnlyViolation):
        open_db(Path("\\\\hlhz\\dfs\\nyfva\\PV\\index.db"), PROD_PV_ROOT)
    with pytest.raises(ReadOnlyViolation):
        open_db(tmp_path / "pv" / "index.db", str(tmp_path / "pv"))


# --------------------------------------------------------------------------
# scan_tree on a synthetic tmp tree
# --------------------------------------------------------------------------


def _write(path: Path, content: bytes = b"data") -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_bytes(content)


def _build_tree(pv: Path) -> dict[str, Path]:
    deal = pv / "Angelo Gordon" / "Accell"
    files = {
        "client_memo": deal / "1.31.2025" / "Client" / "Accell Valuation Memo v2.pdf",
        "analysis": deal / "1.31.2025" / "Analysis" / "Accell 1.31.2025 Report v1.xlsx",
        "archived": deal / "Archive" / "Old Report.pdf",
        "zero": deal / "1.31.2025" / "empty.pdf",
    }
    _write(files["client_memo"], b"memo body")
    _write(files["analysis"], b"model")
    _write(files["archived"], b"superseded content")
    _write(files["zero"], b"")
    return files


@pytest.fixture()
def scan_env(tmp_path: Path):
    pv = tmp_path / "pv"
    pv.mkdir()
    config = Config(pv_root=str(pv), output_dir=tmp_path / "out", db_path=tmp_path / "out" / "scan.db")
    conn = _open_fresh_db(config.db_path, config)
    yield conn, pv, config
    conn.close()


@pytest.mark.skipif(
    sys.platform == "win32" or os.geteuid() == 0,
    reason="symlink + chmod-000 semantics require non-root POSIX",
)
def test_scan_tree_skips_symlinks_and_survives_unreadable_dir(scan_env, tmp_path: Path) -> None:
    conn, pv, config = scan_env
    files = _build_tree(pv)

    outside = tmp_path / "outside"
    _write(outside / "leak.pdf", b"must never be indexed")
    (pv / "Angelo Gordon" / "linked").symlink_to(outside, target_is_directory=True)

    secret = pv / "Angelo Gordon" / "Secret"
    _write(secret / "hidden.pdf", b"unreadable")
    secret.chmod(0o000)
    try:
        stats = scan_tree(conn, pv, config)
    finally:
        secret.chmod(0o755)

    assert stats.files_seen == 4
    assert stats.added == 4
    assert stats.updated == 0
    assert stats.removed == 0
    assert stats.errors == 1
    assert count_files(conn) == 4

    errors = scan_errors_under(conn, str(pv))
    assert len(errors) == 1
    assert errors[0].error_type == "PermissionError"
    assert "Secret" in errors[0].path

    leaked = conn.execute("SELECT COUNT(*) FROM files WHERE file_path LIKE '%leak%'").fetchone()[0]
    assert leaked == 0

    client_record = record_from_row(_row_for(conn, str(files["client_memo"])))
    assert client_record.client == "Angelo Gordon"
    assert client_record.deal == "Accell"
    assert client_record.source_class is SourceClass.client
    assert client_record.as_of_date == date(2025, 1, 31)
    assert client_record.date_folder == "1.31.2025"
    assert not client_record.is_archive

    archived = record_from_row(_row_for(conn, str(files["archived"])))
    assert archived.source_class is SourceClass.archive
    assert archived.is_archive
    assert archived.archive_or_old_flag

    zero = record_from_row(_row_for(conn, str(files["zero"])))
    assert zero.is_zero_byte
    assert zero.size_bytes == 0
    assert not zero.is_cloud_placeholder


def test_scan_tree_incremental_refresh(scan_env) -> None:
    conn, pv, config = scan_env
    files = _build_tree(pv)

    first = scan_tree(conn, pv, config)
    assert (first.files_seen, first.added, first.errors) == (4, 4, 0)
    assert count_files(conn) == 4

    second = scan_tree(conn, pv, config)
    assert second.unchanged == 4
    assert (second.added, second.updated, second.removed) == (0, 0, 0)

    # modify: new content + mtime bumped past second resolution
    target = files["analysis"]
    id_before = _row_for(conn, str(target))["id"]
    target.write_bytes(b"model with much longer content")
    bumped = target.stat().st_mtime + 10
    os.utime(target, (bumped, bumped))
    third = scan_tree(conn, pv, config)
    assert (third.updated, third.unchanged, third.added, third.removed) == (1, 3, 0, 0)
    row = _row_for(conn, str(target))
    assert row["id"] == id_before  # upsert preserves id
    assert row["size_bytes"] == len(b"model with much longer content")

    # delete one
    files["zero"].unlink()
    fourth = scan_tree(conn, pv, config)
    assert (fourth.removed, fourth.unchanged, fourth.files_seen) == (1, 3, 3)
    assert count_files(conn) == 3
    assert str(files["zero"]) not in file_meta_under(conn, str(pv))

    # add one
    new_file = pv / "Angelo Gordon" / "Accell" / "11.30.2024" / "Accell Update.pdf"
    _write(new_file, b"new period")
    fifth = scan_tree(conn, pv, config)
    assert (fifth.added, fifth.unchanged, fifth.removed) == (1, 3, 0)
    assert count_files(conn) == 4
    assert set(file_meta_under(conn, str(pv))) == {
        str(files["client_memo"]),
        str(files["analysis"]),
        str(files["archived"]),
        str(new_file),
    }
    assert as_of_dates_for_deal(conn, "Angelo Gordon", "Accell") == [
        date(2024, 11, 30),
        date(2025, 1, 31),
    ]


def _age_tree(root: Path, seconds: int) -> None:
    """Backdate every directory mtime under `root` so a quick rescan treats
    its leaves as untouched-since-last-scan."""
    past = time.time() - seconds
    for d in [root, *(p for p in root.rglob("*") if p.is_dir())]:
        os.utime(d, (past, past))


def test_scan_tree_quick_rescan_skips_unchanged_leaves(scan_env) -> None:
    conn, pv, config = scan_env
    config.indexer.quick_rescan_margin_seconds = 0  # deterministic cutoff for the test
    files = _build_tree(pv)
    _age_tree(pv, 1000)  # all folder mtimes safely predate the first scan

    first = scan_tree(conn, pv, config, quick=True)  # no baseline yet -> full walk
    assert (first.added, first.files_seen) == (4, 4)
    assert count_files(conn) == 4

    # Nothing changed: the leaf folders (Client/Analysis/Archive) are skipped by
    # mtime, their files still counted as unchanged, and nothing is deleted.
    second = scan_tree(conn, pv, config, quick=True)
    assert (second.added, second.updated, second.removed) == (0, 0, 0)
    assert second.unchanged == 4
    assert second.files_seen == 4
    assert count_files(conn) == 4

    # A genuinely new file bumps its leaf folder's mtime -> that folder is
    # re-listed and the file picked up; the other leaves stay skipped.
    new_file = files["client_memo"].parent / "Accell Valuation Memo v3.pdf"
    _write(new_file, b"newer memo")
    third = scan_tree(conn, pv, config, quick=True)
    assert third.added == 1
    assert count_files(conn) == 5
    assert str(new_file) in file_meta_under(conn, str(pv))

    # A whole new period folder is also caught (its parent deal folder is a
    # non-leaf, always re-listed, so the new subfolder is discovered).
    new_period = pv / "Angelo Gordon" / "Accell" / "11.30.2024" / "Accell Update.pdf"
    _write(new_period, b"new period")
    fourth = scan_tree(conn, pv, config, quick=True)
    assert fourth.added == 1
    assert count_files(conn) == 6


def test_scan_tree_quick_rescan_misses_in_place_overwrite(scan_env) -> None:
    conn, pv, config = scan_env
    config.indexer.quick_rescan_margin_seconds = 0
    files = _build_tree(pv)
    _age_tree(pv, 1000)
    scan_tree(conn, pv, config, quick=True)  # records the baseline

    # Overwrite a file in place under the same name: the file's own size/mtime
    # change but its folder's mtime does not -> keep the folder looking old.
    target = files["archived"]
    target.write_bytes(b"silently replaced content of a very different length")
    bumped = time.time() + 10
    os.utime(target, (bumped, bumped))
    _age_tree(pv, 1000)  # folders still predate the scan; the file mtime stays new

    quick = scan_tree(conn, pv, config, quick=True)
    assert quick.updated == 0  # documented blind spot: in-place edit not seen

    full = scan_tree(conn, pv, config, quick=False)  # a full rescan catches it
    assert full.updated == 1
    assert _row_for(conn, str(target))["size_bytes"] == len(
        b"silently replaced content of a very different length"
    )


def test_fts_stays_in_sync_through_update_and_delete(scan_env) -> None:
    conn, pv, config = scan_env
    files = _build_tree(pv)
    scan_tree(conn, pv, config)
    assert len(fts_candidates(conn, '"valuation memo"', 10)) == 1

    files["client_memo"].unlink()
    renamed = files["client_memo"].with_name("Accell Portfolio Review v3.pdf")
    _write(renamed, b"renamed")
    scan_tree(conn, pv, config)
    assert fts_candidates(conn, '"valuation memo"', 10) == []
    hits = fts_candidates(conn, '"portfolio review"', 10)
    assert [r.file_path for r in hits] == [str(renamed)]
    assert hits[0].contains_memo_keyword


def test_stat_is_cloud_placeholder_flags() -> None:
    assert stat_is_cloud_placeholder(SimpleNamespace(st_file_attributes=0x1000))  # OFFLINE
    assert stat_is_cloud_placeholder(SimpleNamespace(st_file_attributes=0x40000))  # RECALL_ON_OPEN
    assert stat_is_cloud_placeholder(SimpleNamespace(st_file_attributes=0x400000))  # RECALL_ON_DATA_ACCESS
    assert not stat_is_cloud_placeholder(SimpleNamespace(st_file_attributes=0))
    assert not stat_is_cloud_placeholder(os.stat(__file__))  # POSIX stat lacks the attribute


def test_derive_record_rejects_paths_outside_pv_root(prod_config: Config) -> None:
    with pytest.raises(ValueError, match="not under pv_root"):
        derive_record(
            "\\\\otherserver\\share\\file.pdf",
            size_bytes=1,
            modified_time=None,
            config=prod_config,
        )


def test_scan_tree_pause_and_resume(scan_env) -> None:
    """should_stop pauses cooperatively: scanned rows are committed, the
    vanished-paths deletion is SKIPPED (an incomplete walk proves nothing),
    and re-running the same scan completes the index and applies deletions."""
    conn, pv, config = scan_env
    _build_tree(pv)

    # A stale row under the root: a paused scan must NOT delete it.
    stale = derive_record(
        str(pv / "Angelo Gordon" / "Accell" / "gone.pdf"),
        size_bytes=1, modified_time=datetime(2025, 1, 1, 12, 0, 0), config=config,
    )
    insert_records(conn, [stale], batch_size=10)

    calls = {"n": 0}

    def stop_after_two() -> bool:
        calls["n"] += 1
        return calls["n"] > 2

    paused = scan_tree(conn, pv, config, should_stop=stop_after_two)
    assert paused.stopped_early is True
    assert paused.removed == 0
    indexed_while_paused = count_files(conn)
    assert indexed_while_paused >= 1  # partial progress was committed
    assert conn.execute(
        "SELECT COUNT(*) FROM files WHERE file_name = 'gone.pdf'"
    ).fetchone()[0] == 1  # stale row survived the pause

    resumed = scan_tree(conn, pv, config)
    assert resumed.stopped_early is False
    assert resumed.unchanged >= indexed_while_paused - 1  # fast-forward over committed rows
    assert resumed.removed == 1  # complete walk now proves gone.pdf vanished
    assert count_files(conn) == 4


def test_scan_tree_immediate_stop_commits_nothing_and_deletes_nothing(scan_env) -> None:
    conn, pv, config = scan_env
    _build_tree(pv)
    stats = scan_tree(conn, pv, config, should_stop=lambda: True)
    assert stats.stopped_early is True
    assert stats.files_seen == 0 and stats.removed == 0
    assert count_files(conn) == 0
