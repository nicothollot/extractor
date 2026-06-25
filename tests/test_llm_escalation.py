"""D4/D5/D6 escalation tests: payload assembly, quote grounding, merge
policy, tier ladder, budget cap, response cache, and the full pipeline on
the scanned fixture memo. Claude Code is ALWAYS the fake client here."""

from __future__ import annotations

import json
from datetime import datetime
from pathlib import Path

import pytest
from fixtures.docgen import make_scanned_pdf, make_text_pdf
from fixtures.fake_claude import FakeClaudeCodeClient, field_result

from pv_extractor.config import Config
from pv_extractor.extract.engine import load_schema_fields
from pv_extractor.llm.costs import LEDGER_FILENAME, read_ledger, summarize_ledger
from pv_extractor.llm.escalate import (
    LlmSettings,
    process_memos,
    quote_grounding,
    resolve_settings,
)
from pv_extractor.llm.payload import MemoPayload, PayloadPage, assemble_payload, select_pages
from pv_extractor.models import (
    AssetExtraction,
    EscalationField,
    EscalationPlan,
    FieldHit,
    MemoResult,
    PageClass,
    QaStatus,
)

PROJECT_ROOT = Path(__file__).resolve().parent.parent
SCHEMA_FIELDS = load_schema_fields()
BY_HEADER = {f.header: f for f in SCHEMA_FIELDS}

PAGE1 = [
    "Valuation Memo",
    "Fund Name: Test Fund I",
    "Gross IRR: 12.5%",
    "MOIC: 1.4x",
]


# ---------------------------------------------------------------------------
# scaffolding
# ---------------------------------------------------------------------------


def make_config(tmp_path: Path) -> Config:
    config = Config(
        pv_root=str(tmp_path / "not_pv"),
        output_dir=tmp_path / "output",
        db_path=tmp_path / "output" / "pv.db",
    )
    config.llm.models_path = str(PROJECT_ROOT / "config" / "models.yaml")
    # These tests pin the EMBEDDED page-payload path (assembled page text/images
    # in the prompt). direct_document_read (the production default) replaces that
    # with a copy-the-file + Read-it call and is validated separately.
    config.llm.direct_document_read = False
    (tmp_path / "output").mkdir(parents=True, exist_ok=True)
    return config


def escalation_field(header: str, reason: str, pages: list[int] | None = None) -> EscalationField:
    schema_field = BY_HEADER[header]
    return EscalationField(
        field=header, col_index=schema_field.col_index, band=schema_field.band,
        reason=reason, candidate_pages=pages or [1],
    )


def det_hit(header: str, value, confidence: float) -> FieldHit:
    schema_field = BY_HEADER[header]
    return FieldHit(
        field=header, col_index=schema_field.col_index, band=schema_field.band,
        raw_text=str(value), value=value, method="deterministic",
        confidence=confidence, evidence=f"{header}: {value}", page=1,
    )


def make_memo(tmp_path: Path, *, hits: list[FieldHit], plan_fields: list[EscalationField]) -> MemoResult:
    pdf = tmp_path / "docs" / "memo.pdf"
    if not pdf.exists():
        make_text_pdf(pdf, [PAGE1])
    memo = MemoResult(
        memo_id="MEMO_20260611_120000_001", run_id="RUN_TEST", client="TestClient",
        deal="TestDeal", file_path=str(pdf), file_name="memo.pdf",
    )
    memo.assets.append(
        AssetExtraction(
            asset_name="TestDeal", row_memo_id=memo.memo_id, hits=hits,
            qa_status=QaStatus.qa_pass_with_flags,
        )
    )
    memo.escalation = EscalationPlan(
        memo_id=memo.memo_id, confidence_threshold=0.75, fields=plan_fields,
    )
    return memo


def settings(**overrides) -> LlmSettings:
    base = dict(
        enabled=True, mode="auto", manual_model="sonnet", manual_effort="low",
        allow_fable=False, budget_usd=25.0, workers=1, force=False,
    )
    base.update(overrides)
    return LlmSettings(**base)


def run_escalation(config, memos, client, *, llm_settings=None, event_sink=None):
    return process_memos(
        memos, config, llm_settings or settings(), SCHEMA_FIELDS,
        run_id="RUN_TEST", run_dir=Path(config.output_dir) / "RUN_TEST", client=client,
        event_sink=event_sink,
    )


def flags_of(memo: MemoResult) -> list[str]:
    return [flag.description for flag in memo.assets[0].flags]


# ---------------------------------------------------------------------------
# page selection + payload (rules 1 and 8)
# ---------------------------------------------------------------------------


def test_select_pages_targets_candidates_plus_summary():
    fields = [
        escalation_field("Gross IRR %", "below_confidence", pages=[7, 9]),
        escalation_field("Fund Name", "required_empty", pages=[9, 40]),
    ]
    assert select_pages(fields, page_count=30, summary_pages=3, max_pages=20) == [1, 2, 3, 7, 9]


def test_select_pages_caps_at_max_keeping_summary_first():
    fields = [escalation_field("Gross IRR %", "below_confidence", pages=list(range(4, 30)))]
    selected = select_pages(fields, page_count=30, summary_pages=3, max_pages=6)
    assert selected == [1, 2, 3, 4, 5, 6]


def test_select_pages_short_memo_is_whole_memo():
    fields = [escalation_field("Gross IRR %", "below_confidence", pages=[1])]
    assert select_pages(fields, page_count=2, summary_pages=3, max_pages=20) == [1, 2]


def test_assemble_payload_text_pages_inline_no_images(tmp_path):
    config = make_config(tmp_path)
    pdf = tmp_path / "docs" / "memo.pdf"
    make_text_pdf(pdf, [PAGE1], tables_by_page={1: [[["Metric", "Value"], ["Gross IRR %", "12.5%"]]]})
    payload = assemble_payload(
        file_path=str(pdf), fields=[escalation_field("Gross IRR %", "below_confidence")],
        config=config, payload_dir=tmp_path / "output" / "payload",
    )
    assert [p.kind for p in payload.pages] == ["text"]
    assert payload.image_count == 0 and payload.ocr_hostile is False
    assert "Fund Name: Test Fund I" in payload.dynamic_prompt
    assert "| Metric | Value |" in payload.dynamic_prompt  # tables ride as pipe tables
    assert (tmp_path / "output" / "payload" / "manifest.json").exists()
    assert (tmp_path / "output" / "payload" / "pages" / "page_001.txt").exists()
    assert payload.payload_hash


def test_direct_document_read_copies_source_and_builds_read_instruction(tmp_path):
    """direct_document_read: the source file is copied into the call dir and the
    prompt points the model at it (Read it directly) instead of embedding the
    rendered page payload."""
    config = make_config(tmp_path)
    config.llm.direct_document_read = True
    pdf = tmp_path / "docs" / "memo.pdf"
    make_text_pdf(pdf, [PAGE1])
    payload = assemble_payload(
        file_path=str(pdf), fields=[escalation_field("Fund Name", "required")],
        config=config, payload_dir=tmp_path / "output" / "payload",
    )
    assert payload.source_documents, "the source document should be copied in"
    _doc_id, rel = payload.source_documents[0]
    assert (payload.directory / rel).exists()
    instruction = payload.read_instruction()
    assert "SOURCE DOCUMENTS" in instruction
    assert rel in instruction and "Read tool" in instruction
    # grounding text is still kept even though it is not embedded in the prompt
    assert payload.page_texts


def test_cleanly_ocrd_scanned_page_travels_as_text_not_image(tmp_path):
    """prefer_ocr_text_over_image (default): a scanned page that OCRs above the
    confidence floor is sent as OCR TEXT — no slow Read-tool/vision call — and
    the payload is no longer OCR-hostile (AUTO can stay on sonnet)."""
    config = make_config(tmp_path)
    pdf = tmp_path / "docs" / "scan.pdf"
    make_scanned_pdf(pdf, [["Valuation Memo", "Gross IRR: 12.5%", "MOIC: 1.4x", "Fund Name: Test Fund I"]])
    fields = [escalation_field("Gross IRR %", "below_confidence")]

    config.llm.prefer_ocr_text_over_image = True
    config.llm.ocr_text_min_confidence = 0.5  # synthetic scan OCRs well above this
    text_payload = assemble_payload(
        file_path=str(pdf), fields=fields, config=config,
        payload_dir=tmp_path / "out" / "text",
    )
    assert [p.kind for p in text_payload.pages] == ["text"]
    assert text_payload.image_count == 0
    assert text_payload.ocr_hostile is False
    assert "Read tool" not in text_payload.dynamic_prompt  # OCR text inline, not an image ref

    config.llm.prefer_ocr_text_over_image = False
    image_payload = assemble_payload(
        file_path=str(pdf), fields=fields, config=config,
        payload_dir=tmp_path / "out" / "image",
    )
    assert [p.kind for p in image_payload.pages] == ["image"]
    assert image_payload.image_count == 1 and image_payload.ocr_hostile is True


# ---------------------------------------------------------------------------
# quote grounding (rule 5)
# ---------------------------------------------------------------------------


def _payload_with(page_texts: dict[int, str], image_pages: set[int] = frozenset()) -> MemoPayload:
    payload = MemoPayload(directory=Path("."))
    payload.page_texts = page_texts
    payload.pages = [
        PayloadPage(
            number=n, page_class=PageClass.SCANNED if n in image_pages else PageClass.TEXT,
            kind="image" if n in image_pages else "text", rel_path="x", sha256="",
        )
        for n in page_texts
    ]
    return payload


def test_quote_grounding_exact_and_normalized():
    payload = _payload_with({1: "Fund  Name:\nTest Fund I"})
    assert quote_grounding("fund name: test fund i", 1, payload, 85) == "grounded"
    assert quote_grounding("Test Fund I", 1, payload, 85) == "grounded"


def test_quote_grounding_rejects_fabrications_and_wrong_pages():
    payload = _payload_with({1: "Gross IRR: 12.5%"})
    assert quote_grounding("Gross IRR: 99.9%", 1, payload, 85) == "ungrounded"
    assert quote_grounding("Gross IRR: 12.5%", 2, payload, 85) == "ungrounded"
    assert quote_grounding("", 1, payload, 85) == "ungrounded"


def test_quote_grounding_is_fuzzy_on_image_pages_only():
    noisy = "Gros5 IRR : 12.5 %"  # OCR noise
    payload = _payload_with({1: noisy}, image_pages={1})
    assert quote_grounding("Gross IRR: 12.5%", 1, payload, 85) == "grounded"
    payload = _payload_with({1: noisy})
    assert quote_grounding("Gross IRR: 12.5%", 1, payload, 85) == "ungrounded"


def test_quote_grounding_unverifiable_without_local_text():
    payload = _payload_with({4: ""}, image_pages={4})
    assert quote_grounding("anything", 4, payload, 85) == "unverifiable"


# ---------------------------------------------------------------------------
# merge policy + tier ladder
# ---------------------------------------------------------------------------


def test_fill_empty_field_and_never_overwrite_confident_value(tmp_path):
    config = make_config(tmp_path)
    memo = make_memo(
        tmp_path,
        hits=[det_hit("Gross IRR %", 11.0, confidence=0.95)],  # confident: untouchable
        plan_fields=[
            escalation_field("Fund Name", "required_empty"),
            escalation_field("Gross IRR %", "below_confidence"),
        ],
    )
    fake = FakeClaudeCodeClient({
        "Fund Name": field_result("Test Fund I", page=1, quote="Fund Name: Test Fund I"),
        "Gross IRR %": field_result(12.5, unit="percent", page=1, quote="Gross IRR: 12.5%"),
    })
    summary = run_escalation(config, [memo], fake)

    fund = next(h for h in memo.assets[0].hits if h.field == "Fund Name")
    assert fund.value == "Test Fund I"
    assert fund.method == "llm:claude:sonnet:medium"  # text memo -> AUTO tier 0
    assert fund.evidence == "Fund Name: Test Fund I" and fund.page == 1
    assert fund.confidence == pytest.approx(0.85)

    irr = next(h for h in memo.assets[0].hits if h.field == "Gross IRR %")
    assert irr.value == 11.0 and irr.method == "deterministic"  # never overwritten

    plan = memo.escalation
    assert plan.merged_fields == ["Fund Name"]
    assert any("rejected:protected" in line for line in plan.merge_log)
    assert plan.status == "llm_partial"
    assert any(d.startswith("LLM_UNCONFIRMED: Gross IRR") for d in flags_of(memo))
    tiers = [(a.model_alias, a.effort) for a in plan.attempts]
    assert tiers == [("sonnet", "medium")]
    assert summary.attempts == len(plan.attempts) and summary.total_cost_usd > 0
    ledger = read_ledger(Path(config.output_dir) / "RUN_TEST" / "llm" / LEDGER_FILENAME)
    assert len(ledger) == len(plan.attempts) and all(e["cost_source"] == "estimated" for e in ledger)


def test_overwrite_below_threshold_keeps_old_value_as_conflict(tmp_path):
    config = make_config(tmp_path)
    memo = make_memo(
        tmp_path,
        hits=[det_hit("Gross IRR %", 11.0, confidence=0.50)],
        plan_fields=[escalation_field("Gross IRR %", "below_confidence")],
    )
    fake = FakeClaudeCodeClient(
        {"Gross IRR %": field_result(12.5, unit="percent", page=1, quote="Gross IRR: 12.5%")}
    )
    run_escalation(config, [memo], fake)

    irr = next(h for h in memo.assets[0].hits if h.field == "Gross IRR %")
    assert irr.value == 12.5 and irr.method == "llm:claude:sonnet:medium"
    assert irr.conflicts and irr.conflicts[-1].value == 11.0  # audit keeps the loser
    plan = memo.escalation
    assert plan.status == "llm_completed"
    assert any("overwrote below-threshold" in line for line in plan.merge_log)
    assert len(plan.attempts) == 1  # resolved at tier 0, no retry


def test_ungrounded_value_is_ineligible_by_default(tmp_path):
    """Default arbitration requires grounded evidence: an ungrounded value is
    retained as an unresolved review item rather than filling the field."""
    config = make_config(tmp_path)
    memo = make_memo(tmp_path, hits=[], plan_fields=[escalation_field("Fund Name", "required_empty")])
    fake = FakeClaudeCodeClient(
        {"Fund Name": field_result("Fake Fund", page=1, quote="Fund Name: Fake Fund LP")}
    )
    run_escalation(config, [memo], fake)

    assert next((h for h in memo.assets[0].hits if h.field == "Fund Name"), None) is None
    assert any(d.startswith("LLM_UNRESOLVED: Fund Name") for d in flags_of(memo))


def test_ungrounded_value_discarded_when_surfacing_disabled(tmp_path):
    """surface_ungrounded_values=False restores the strict behavior: the value
    is discarded and the required field surfaces NOT_EXTRACTABLE."""
    config = make_config(tmp_path)
    config.llm.surface_ungrounded_values = False
    memo = make_memo(tmp_path, hits=[], plan_fields=[escalation_field("Fund Name", "required_empty")])
    fake = FakeClaudeCodeClient(
        {"Fund Name": field_result("Fake Fund", page=1, quote="Fund Name: Fake Fund LP")}
    )
    run_escalation(config, [memo], fake)

    assert not any(h.field == "Fund Name" for h in memo.assets[0].hits)  # value discarded
    assert any(d.startswith("UNGROUNDED_LLM_VALUE: Fund Name") for d in flags_of(memo))
    assert "Fund Name" in memo.escalation.not_extractable
    assert any(d.startswith("NOT_EXTRACTABLE: Fund Name") for d in flags_of(memo))


def test_wholesale_call_failure_emits_one_clear_error_not_per_field_noise(tmp_path):
    """When EVERY Claude Code call fails, emit ONE actionable LLM_PASS_FAILED
    flag carrying the CLI's real error — not an identical 'no value' flag per
    field (which buries the cause under hundreds of rows)."""
    config = make_config(tmp_path)
    memo = make_memo(
        tmp_path, hits=[],
        plan_fields=[escalation_field(h, "required_empty") for h in ("Fund Name", "MOIC")],
    )
    fake = FakeClaudeCodeClient({}, behaviors=["exit", "exit", "exit", "exit"])  # all calls fail
    run_escalation(config, [memo], fake)

    flags = flags_of(memo)
    pass_failed = [d for d in flags if d.startswith("LLM_PASS_FAILED")]
    assert len(pass_failed) == 1
    assert "exit 3" in pass_failed[0]  # the CLI's real error rides along
    assert not any(d.startswith("NOT_EXTRACTABLE") for d in flags)  # no per-field noise
    assert memo.escalation.status == "llm_failed"


def test_malformed_json_does_not_trigger_routine_retry(tmp_path):
    config = make_config(tmp_path)
    memo = make_memo(tmp_path, hits=[], plan_fields=[escalation_field("Fund Name", "required_empty")])
    fake = FakeClaudeCodeClient(
        {"Fund Name": field_result("Test Fund I", page=1, quote="Fund Name: Test Fund I")},
        behaviors=["malformed", "ok"],
    )
    run_escalation(config, [memo], fake)

    plan = memo.escalation
    assert plan.attempts[0].error and "non-JSON" in plan.attempts[0].error
    assert len(plan.attempts) == 1
    assert next((h for h in memo.assets[0].hits if h.field == "Fund Name"), None) is None
    assert plan.status == "llm_failed"


def test_not_found_is_resolved_no_retry_no_flag(tmp_path):
    """Default (retry_not_found=False): a field the model marks not_found is a
    CONFIRMED ABSENCE — no expensive-tier re-ask, no NOT_EXTRACTABLE flag."""
    config = make_config(tmp_path)
    memo = make_memo(
        tmp_path,
        hits=[det_hit("Gross IRR %", 11.0, confidence=0.50)],
        plan_fields=[
            escalation_field("Fund Name", "required_empty"),
            escalation_field("Gross IRR %", "below_confidence"),
        ],
    )
    fake = FakeClaudeCodeClient({})  # everything not_found
    run_escalation(config, [memo], fake)

    plan = memo.escalation
    assert plan.merged_fields == []
    # One sonnet call; both fields report not_found and are resolved, so the
    # opus retry tier is never invoked (no opus call, 2 not_found not 4).
    assert len(fake.calls) == 1
    assert all(c["model"] != "opus" for c in fake.calls)
    assert sum(a.fields_not_found for a in plan.attempts) == 2
    # confirmed-absent fields are NOT flagged as extraction failures
    assert plan.not_extractable == []
    assert not any(d.startswith("NOT_EXTRACTABLE") for d in flags_of(memo))
    assert not any(d.startswith("LLM_UNCONFIRMED") for d in flags_of(memo))
    irr = next(h for h in memo.assets[0].hits if h.field == "Gross IRR %")
    assert irr.value == 11.0  # untouched


def test_retry_not_found_true_no_longer_creates_model_ladder(tmp_path):
    """Legacy retry_not_found no longer creates a stronger model ladder; repair
    is controlled by candidate_arbitration.repair_policy."""
    config = make_config(tmp_path)
    config.llm.retry_not_found = True
    memo = make_memo(
        tmp_path,
        hits=[det_hit("Gross IRR %", 11.0, confidence=0.50)],
        plan_fields=[
            escalation_field("Fund Name", "required_empty"),
            escalation_field("Gross IRR %", "below_confidence"),
        ],
    )
    fake = FakeClaudeCodeClient({})  # everything not_found
    run_escalation(config, [memo], fake)

    plan = memo.escalation
    assert plan.merged_fields == []
    assert all(c["model"] != "opus" for c in fake.calls)
    assert sum(a.fields_not_found for a in plan.attempts) == 2
    assert plan.not_extractable == ["Fund Name"]
    assert any(d.startswith("NOT_EXTRACTABLE: Fund Name") for d in flags_of(memo))
    assert any(d.startswith("LLM_UNCONFIRMED: Gross IRR") for d in flags_of(memo))
    irr = next(h for h in memo.assets[0].hits if h.field == "Gross IRR %")
    assert irr.value == 11.0  # untouched


def test_budget_cap_defers_without_any_call(tmp_path):
    config = make_config(tmp_path)
    memo = make_memo(tmp_path, hits=[], plan_fields=[escalation_field("Fund Name", "required_empty")])
    fake = FakeClaudeCodeClient(
        {"Fund Name": field_result("Test Fund I", page=1, quote="Fund Name: Test Fund I")}
    )
    summary = run_escalation(config, [memo], fake, llm_settings=settings(budget_usd=0.0))

    assert fake.calls == []  # never submitted
    plan = memo.escalation
    assert plan.status == "llm_deferred_budget" and plan.attempts == []
    assert any(d.startswith("LLM_DEFERRED") for d in flags_of(memo))
    assert summary.memos_deferred == 1


def test_response_cache_prevents_repaying_and_force_bypasses(tmp_path):
    config = make_config(tmp_path)
    values = {"Fund Name": field_result("Test Fund I", page=1, quote="Fund Name: Test Fund I")}
    plan_fields = [escalation_field("Fund Name", "required_empty")]

    fake1 = FakeClaudeCodeClient(values)
    memo1 = make_memo(tmp_path, hits=[], plan_fields=list(plan_fields))
    run_escalation(config, [memo1], fake1)
    assert len(fake1.calls) == 1 and memo1.escalation.attempts[0].from_cache is False

    fake2 = FakeClaudeCodeClient(values)
    memo2 = make_memo(tmp_path, hits=[], plan_fields=list(plan_fields))
    run_escalation(config, [memo2], fake2)
    assert fake2.calls == []  # rule 10: unchanged memo never re-runs
    attempt = memo2.escalation.attempts[0]
    assert attempt.from_cache is True and attempt.cost_usd == 0.0
    fund = next(h for h in memo2.assets[0].hits if h.field == "Fund Name")
    assert fund.value == "Test Fund I"  # merged from the cached response

    fake3 = FakeClaudeCodeClient(values)
    memo3 = make_memo(tmp_path, hits=[], plan_fields=list(plan_fields))
    run_escalation(config, [memo3], fake3, llm_settings=settings(force=True))
    assert len(fake3.calls) == 1  # --force-llm bypasses the cache


def test_llm_activity_events_capture_call_lifecycle_without_page_text(tmp_path):
    config = make_config(tmp_path)
    memo = make_memo(tmp_path, hits=[], plan_fields=[escalation_field("Fund Name", "required_empty")])
    fake = FakeClaudeCodeClient(
        {"Fund Name": field_result("Test Fund I", page=1, quote="Fund Name: Test Fund I")}
    )
    events: list[tuple[str, dict[str, object]]] = []

    run_escalation(config, [memo], fake, event_sink=lambda event, payload: events.append((event, payload)))

    activity = [payload for event, payload in events if event == "llm_activity"]
    assert [payload["status"] for payload in activity] == ["started", "finished"]
    started = activity[0]
    assert started["provider"] == "claude"
    assert started["model"] == "sonnet"
    assert started["fields_requested"] == 1
    assert started["selected_pages"] == [1]
    assert started["documents"] == [
        {"document_id": "D01", "name": "memo.pdf", "path": str(tmp_path / "docs" / "memo.pdf")}
    ]
    prompt_path = Path(str(started["prompt_path"]))
    assert prompt_path.exists()
    preview = str(started["prompt_preview"])
    assert "Fields to extract" in preview
    assert "document/page payload omitted" in preview
    assert "Fund Name: Test Fund I" not in preview
    assert "Fund Name: Test Fund I" in prompt_path.read_text(encoding="utf-8")
    assert activity[-1]["status"] == "finished"
    assert activity[-1]["session_id"] == "sess-fake-001"


def test_auth_failure_fails_plans_with_login_instruction(tmp_path):
    config = make_config(tmp_path)
    memo = make_memo(tmp_path, hits=[], plan_fields=[escalation_field("Fund Name", "required_empty")])
    fake = FakeClaudeCodeClient(auth_ok=False)
    summary = run_escalation(config, [memo], fake)

    assert fake.calls == []
    assert memo.escalation.status == "llm_failed"
    assert "claude auth login" in summary.detail
    assert any("claude auth login" in d for d in flags_of(memo))


def test_manual_mode_forces_one_model_for_everything(tmp_path):
    config = make_config(tmp_path)
    memo = make_memo(
        tmp_path,
        hits=[],
        plan_fields=[
            escalation_field("Fund Name", "required_empty"),
            escalation_field("Gross IRR %", "below_confidence"),
        ],
    )
    fake = FakeClaudeCodeClient(
        {"Fund Name": field_result("Test Fund I", page=1, quote="Fund Name: Test Fund I")}
    )
    run_escalation(
        config, [memo], fake,
        llm_settings=settings(mode="manual", manual_model="sonnet", manual_effort="low"),
    )
    plan = memo.escalation
    # Manual mode forces the one model+effort for every call and never escalates
    # to the opus retry tier. Adaptive batching merges FUND + RETURNS (both on
    # page 1) into a single sonnet/low call.
    assert plan.attempts
    assert all((a.model_alias, a.effort) == ("sonnet", "low") for a in plan.attempts)
    assert all(a.model_alias == "sonnet" for a in plan.attempts)
    assert len(fake.calls) == len(plan.attempts)  # merged: one call (no opus retry in manual mode)


def test_resolve_settings_cli_semantics(default_config):
    s = resolve_settings(default_config)
    assert s.enabled and s.mode == "auto" and s.budget_usd == 25.0
    assert resolve_settings(default_config, no_llm=True).enabled is False
    s = resolve_settings(default_config, model="haiku", effort="low")
    assert s.mode == "single_model" and s.manual_model == "haiku"
    assert resolve_settings(default_config, budget=5.0).budget_usd == 5.0


# ---------------------------------------------------------------------------
# full pipeline on the scanned fixture memo (D6)
# ---------------------------------------------------------------------------


def test_full_pipeline_scanned_memo_escalates_merges_and_ledgers(phase2_env):
    """Deterministic OCR scores below threshold -> EscalationPlan -> fake
    Claude Code -> merged row + audit record + cost ledger entries."""
    from openpyxl import load_workbook

    from pv_extractor.run import run
    from pv_extractor.write import RUN_LOG_COLUMNS

    config = phase2_env
    # Keep exercising the image/vision (opus) path here: force the scanned pages
    # to travel as images even though they OCR cleanly (the text-downgrade path
    # has its own test below).
    config.llm.prefer_ocr_text_over_image = False
    fake = FakeClaudeCodeClient({
        "Gross IRR %": field_result(12.5, unit="percent", page=1, quote="Gross IRR: 12.5%"),
        # fabricated quote: must be rejected by quote-grounding
        "MOIC": field_result(9.9, unit="x", page=1, quote="Quarterly dividend of $4.21 per share"),
    })
    report = run(
        config, scope="deal", client="Angeles Investments", deal="Andover Storage",
        period="2026-03-31", force=True, now=datetime(2026, 6, 11, 13, 0, 0),
        llm_settings=resolve_settings(config), llm_client=fake,
    )

    assert report.coverage[0].status == "FOUND"
    memo = report.memos[0]
    plan = memo.escalation
    escalated = {f.field for f in plan.fields}
    assert {"Gross IRR %", "MOIC"} <= escalated  # OCR-page hits sit below threshold

    # A one-document deal is one primary request; AUTO stays on the configured
    # normal extraction model unless profile metrics justify Opus.
    assert fake.calls and fake.calls[0]["model"] == "sonnet" and fake.calls[0]["effort"] == "medium"

    # merged row: the grounded LLM value replaced the low-confidence OCR hit
    irr = next(h for h in memo.assets[0].hits if h.field == "Gross IRR %")
    assert irr.value == 12.5
    assert irr.method == "llm:claude:sonnet:medium"
    assert irr.evidence == "Gross IRR: 12.5%"
    assert "Gross IRR %" in plan.merged_fields

    # fabricated quote: value discarded, unresolved review item raised
    moic = next(h for h in memo.assets[0].hits if h.field == "MOIC")
    assert moic.method != "llm:claude:opus:high" or moic.value != 9.9
    assert any(
        f.description.startswith("LLM_UNRESOLVED: MOIC") and f.reviewer_attention
        for f in memo.assets[0].flags
    )

    # attempts + run summary + ledger
    assert plan.status in ("llm_partial", "llm_completed")
    assert len(plan.attempts) >= 1
    assert report.llm is not None and report.llm.executed
    assert report.llm.session_labels and report.llm.total_cost_usd > 0
    ledger_entries = read_ledger(report.run_dir / "llm" / LEDGER_FILENAME)
    assert len(ledger_entries) == len(plan.attempts)
    summary = summarize_ledger(ledger_entries)
    assert summary["memos"] == 1 and summary["total_usd"] > 0

    # audit record carries the attempts and the merge log
    audit = json.loads((report.run_dir / "audit" / f"{memo.memo_id}.json").read_text("utf-8"))
    assert audit["escalation"]["status"] == plan.status
    assert audit["escalation"]["attempts"][0]["job_id"].startswith(f"pv-{report.run_id}-")
    assert audit["escalation"]["merge_log"]

    # workbook: merged value landed in the Index row; sessions in the Run Log
    workbook = load_workbook(report.workbook_path, read_only=True)
    index = workbook["Index"]
    irr_col = BY_HEADER["Gross IRR %"].col_index
    row = next(
        r for r in range(4, index.max_row + 1)
        if index.cell(row=r, column=1).value == memo.memo_id
    )
    assert index.cell(row=row, column=irr_col).value == 12.5
    run_log = workbook["Run Log"]
    sessions_col = RUN_LOG_COLUMNS.index("Batch Sessions") + 1
    sessions_cell = run_log.cell(row=run_log.max_row, column=sessions_col).value
    assert sessions_cell and f"pv-{report.run_id}-{memo.memo_id}-w" in sessions_cell
    workbook.close()


def test_full_pipeline_no_llm_settings_is_pure_phase2(phase2_env):
    from pv_extractor.run import run

    report = run(
        phase2_env, scope="deal", client="Angeles Investments", deal="Andover Storage",
        period="2026-03-31", force=True, now=datetime(2026, 6, 11, 13, 30, 0),
    )
    memo = report.memos[0]
    assert report.llm is None
    assert memo.escalation.attempts == []
    assert memo.escalation.status == "llm_fallback_disabled"
    assert not any(h.method.startswith("llm:") for h in memo.assets[0].hits)


# ---------------------------------------------------------------------------
# one call per deal (default): combine all the deal's documents into ONE call
# ---------------------------------------------------------------------------


def test_assemble_deal_payload_combines_documents_under_one_page_index(tmp_path):
    """A deal's documents are concatenated into ONE payload under a global page
    index; each block is labelled with its source document + that document's own
    page number, and quote-grounding keys off the global index."""
    from pv_extractor.llm.payload import assemble_deal_payload

    config = make_config(tmp_path)
    doc_a = tmp_path / "docs" / "memo_a.pdf"
    doc_b = tmp_path / "docs" / "ic_b.pdf"
    make_text_pdf(doc_a, [["Valuation Memo A", "Gross IRR: 12.5%"]])
    make_text_pdf(doc_b, [["IC Memo B", "MOIC: 1.4x", "Sector: Storage"]])

    payload = assemble_deal_payload(
        files=[(str(doc_a), "Memo A"), (str(doc_b), "IC Memo B")],
        fields=[escalation_field("Gross IRR %", "below_confidence")],
        config=config, payload_dir=tmp_path / "output" / "deal_payload",
    )

    # one global page per single-page doc, numbered across the deal
    assert [p.number for p in payload.pages] == [1, 2]
    # each block names its source document + that document's own page number
    assert "Memo A, document page 1" in payload.dynamic_prompt
    assert "IC Memo B, document page 1" in payload.dynamic_prompt
    # BOTH documents' content is in the single payload
    assert "Gross IRR: 12.5%" in payload.dynamic_prompt
    assert "MOIC: 1.4x" in payload.dynamic_prompt
    # grounding resolves against the GLOBAL page index (MOIC is on global page 2)
    assert quote_grounding("MOIC: 1.4x", 2, payload, 85) == "grounded"
    assert quote_grounding("Gross IRR: 12.5%", 1, payload, 85) == "grounded"
    assert payload.payload_hash
    # per-document payload files are namespaced so two docs never collide
    assert (tmp_path / "output" / "deal_payload" / "pages" / "doc00_page_001.txt").exists()
    assert (tmp_path / "output" / "deal_payload" / "pages" / "doc01_page_001.txt").exists()


def test_one_call_per_deal_extracts_all_documents_in_one_call(tmp_path):
    """process_deals makes ONE Claude Code call per deal over the combined
    payload of all its documents, and merges values from EVERY document onto the
    primary row — no per-band, per-document multi-call fan-out."""
    from pv_extractor.llm.escalate import DealGroup, process_deals

    config = make_config(tmp_path)
    primary_pdf = tmp_path / "docs" / "primary.pdf"
    support_pdf = tmp_path / "docs" / "support.pdf"
    make_text_pdf(primary_pdf, [["Valuation Memo", "Fund Name: Test Fund I"]])
    make_text_pdf(support_pdf, [["IC Memo", "MOIC: 1.4x"]])

    primary = MemoResult(
        memo_id="MEMO_A", run_id="RUN_TEST", client="C", deal="D",
        file_path=str(primary_pdf), file_name="primary.pdf", page_count=1,
    )
    primary.assets.append(
        AssetExtraction(asset_name="D", row_memo_id="MEMO_A", hits=[],
                        qa_status=QaStatus.qa_pass_with_flags)
    )
    primary.escalation = EscalationPlan(
        memo_id="MEMO_A", confidence_threshold=0.75,
        fields=[escalation_field("Gross IRR %", "force_llm_assist"),
                escalation_field("MOIC", "force_llm_assist", pages=[2])],
    )
    support = MemoResult(
        memo_id="MEMO_B", run_id="RUN_TEST", client="C", deal="D",
        file_path=str(support_pdf), file_name="support.pdf", page_count=1,
    )
    support.escalation = EscalationPlan(memo_id="MEMO_B", confidence_threshold=0.75, fields=[])
    group = DealGroup(
        primary=primary, members=[primary, support],
        files=[(str(primary_pdf), "primary.pdf"), (str(support_pdf), "support.pdf")],
    )

    # MOIC's grounding quote lives in the SUPPORT document (global page 2).
    fake = FakeClaudeCodeClient({
        "Gross IRR %": field_result(12.5, unit="percent", page=1, quote="Fund Name: Test Fund I"),
        "MOIC": field_result(1.4, unit="x", page=2, quote="MOIC: 1.4x"),
    })
    summary = process_deals(
        [group], config, settings(), SCHEMA_FIELDS,
        run_id="RUN_TEST", run_dir=Path(config.output_dir) / "RUN_TEST", client=fake,
    )

    # ONE call covered the whole deal (both documents, all fields).
    assert len(fake.calls) == 1
    assert len(primary.escalation.attempts) == 1
    assert summary.executed and summary.attempts == 1
    # values from BOTH documents landed on the primary row, grounded.
    hits = {h.field: h for h in primary.assets[0].hits}
    assert hits["Gross IRR %"].value == 12.5
    assert hits["Gross IRR %"].method.startswith("llm:claude:")
    assert hits["MOIC"].value == 1.4 and hits["MOIC"].page == 2


# ---------------------------------------------------------------------------
# adaptive page-locality batching (efficiency)
# ---------------------------------------------------------------------------

def _unit(label, pages, n, *, must_try=False, priority=True, relevance=1.0):
    from pv_extractor.llm.escalate import _FieldGroup
    fields = [
        EscalationField(field=f"{label}-{i}", col_index=i, band=label,
                        reason="required_empty" if must_try else "force_llm_assist",
                        candidate_pages=list(pages))
        for i in range(n)
    ]
    return _FieldGroup(label=label, fields=fields, pages=sorted(pages), relevance=relevance,
                       has_evidence=True, ocr_hostile=False, priority=priority, must_try=must_try)


def test_pack_merges_same_page_bands_into_few_calls():
    """Many bands all on pages 1-3 collapse into ceil(total/cap) calls, not one
    call per band — the small-document win."""
    from pv_extractor.config import Config
    from pv_extractor.llm.escalate import _pack_by_pages

    config = Config()
    config.llm.max_fields_per_call = 50
    units = [_unit(f"BAND{b}", [1, 2, 3], 20) for b in range(9)]  # 9 bands x 20 = 180 fields
    groups = _pack_by_pages(units, config)
    # 180 fields over one page-set -> ceil(180/50) = 4 calls, not 9
    assert len(groups) == 4
    assert all(g.pages == [1, 2, 3] for g in groups)
    assert sum(len(g.fields) for g in groups) == 180


def test_pack_keeps_distant_pages_separate():
    """Bands whose page-sets can't be unioned under the page budget stay in
    separate calls — a large document still fans out by where its data lives."""
    from pv_extractor.config import Config
    from pv_extractor.llm.escalate import _pack_by_pages

    config = Config()
    config.llm.max_fields_per_call = 50
    config.llm.adaptive_max_pages_per_call = 4
    a = _unit("A", [1, 2], 5)
    b = _unit("B", [40, 41, 42, 43, 44], 5)  # union with A = 7 pages > 4 -> can't merge
    groups = _pack_by_pages([a, b], config)
    assert len(groups) == 2
    page_sets = sorted(tuple(g.pages) for g in groups)
    assert page_sets == [(1, 2), (40, 41, 42, 43, 44)]
