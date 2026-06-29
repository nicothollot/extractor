"""D6/D8 writer tests: template copy semantics, header-drift abort, append
by column index, idempotent cumulative re-runs (cache hit, no duplicate
rows), joint-vehicle rows, Review Flags dedupe, Run Log row."""

from __future__ import annotations

import json
from datetime import datetime
from pathlib import Path

import openpyxl
import pytest

from pv_extractor.run import run
from pv_extractor.write import HeaderDriftError, RUN_LOG_COLUMNS, WorkbookWriter, copy_template

REFERENCE_TEMPLATE = Path(__file__).resolve().parent.parent / "reference" / "master_index_v4.xlsx"


def _schema_fields(project_root: Path):
    from pv_extractor.models import SchemaField

    doc = json.loads((project_root / "schema" / "master_schema.json").read_text(encoding="utf-8"))
    return [SchemaField.model_validate(field) for field in doc["fields"]]


def _index_rows(path: Path) -> list[str]:
    workbook = openpyxl.load_workbook(path, read_only=True)
    sheet = workbook["Index"]
    memo_ids = [
        str(row[0]) for row in sheet.iter_rows(min_row=4, max_col=1, values_only=True) if row[0]
    ]
    workbook.close()
    return memo_ids


def _flag_rows(path: Path) -> list[tuple]:
    workbook = openpyxl.load_workbook(path, read_only=True)
    sheet = workbook["Review Flags"]
    rows = [row for row in sheet.iter_rows(min_row=2, values_only=True) if row[0]]
    workbook.close()
    return rows


# --------------------------------------------------------------------------
# header drift => hard abort before any write
# --------------------------------------------------------------------------


def test_header_drift_template_aborts(project_root: Path, tmp_path: Path) -> None:
    drifted = tmp_path / "drifted.xlsx"
    copy_template(REFERENCE_TEMPLATE, drifted, pv_root=str(tmp_path / "fake_pv"))
    workbook = openpyxl.load_workbook(drifted)
    workbook["Index"].cell(row=2, column=45).value = "Implied EV (renamed by intern)"
    workbook.save(drifted)

    with pytest.raises(HeaderDriftError, match="column 45"):
        WorkbookWriter(drifted, _schema_fields(project_root), pv_root=str(tmp_path / "fake_pv"))


def test_non_master_template_is_custom_and_needs_llm(
    phase2_env, project_root: Path, tmp_path: Path
) -> None:
    """A workbook whose headers no longer match the committed master schema is
    now treated as a CUSTOM reference (extracted LLM-first), not a hard abort.
    With LLM assist disabled the run refuses with a clear ValueError rather
    than the legacy HeaderDriftError."""
    drifted = tmp_path / "drifted.xlsx"
    copy_template(REFERENCE_TEMPLATE, drifted, pv_root=phase2_env.pv_root)
    workbook = openpyxl.load_workbook(drifted)
    workbook["Index"].cell(row=2, column=1).value = "Memo Identifier"
    workbook.save(drifted)

    with pytest.raises(ValueError, match="custom reference workbook"):
        run(
            phase2_env, scope="deal", client="Angelo Gordon", deal="Accell",
            period="2025-01-31", template=drifted, now=datetime(2026, 6, 12, 9, 0, 0),
        )


# --------------------------------------------------------------------------
# idempotent cumulative runs + flag dedupe (cache hit, no duplicate rows)
# --------------------------------------------------------------------------


def test_run_twice_on_same_output_is_idempotent(phase2_env) -> None:
    # Hyperoptic carries a review flag, so this also proves flag dedupe.
    first = run(
        phase2_env, scope="deal", client="Apollo Global Management", deal="Hyperoptic",
        period="Q1 2026", now=datetime(2026, 6, 12, 10, 0, 0),
    )
    assert first.rows_added == 1 and first.flags_added >= 1
    rows_before = _index_rows(first.workbook_path)
    flags_before = _flag_rows(first.workbook_path)

    second = run(
        phase2_env, scope="deal", client="Apollo Global Management", deal="Hyperoptic",
        period="Q1 2026", template=first.workbook_path, now=datetime(2026, 6, 12, 10, 30, 0),
    )
    assert second.cache_hits == 1
    assert second.rows_added == 0  # no duplicate Index rows
    assert second.flags_added == 0  # (memo_id, description) dedupe held
    assert _index_rows(second.workbook_path) == rows_before
    assert len(_flag_rows(second.workbook_path)) == len(flags_before)
    # the cached memo is reported, marked as such
    assert second.memos[0].from_cache is True
    assert second.memos[0].memo_id == first.memos[0].memo_id


def test_force_bypasses_cache(phase2_env) -> None:
    first = run(
        phase2_env, scope="deal", client="Angelo Gordon", deal="Digital Edge",
        period="Q1 2026", now=datetime(2026, 6, 12, 11, 0, 0),
    )
    forced = run(
        phase2_env, scope="deal", client="Angelo Gordon", deal="Digital Edge",
        period="Q1 2026", force=True, now=datetime(2026, 6, 12, 11, 30, 0),
    )
    assert first.cache_hits in (0, 1)  # may hit cache from earlier sessions
    assert forced.cache_hits == 0
    assert forced.memos[0].from_cache is False


# --------------------------------------------------------------------------
# joint-vehicle memo => two rows, suffixed memo ids
# --------------------------------------------------------------------------


def test_joint_vehicle_memo_writes_two_rows(phase2_env) -> None:
    report = run(
        phase2_env, scope="deal", client="Apollo Global Management", deal="AIOF II ANRP III",
        period="Q1 2026", force=True, now=datetime(2026, 6, 12, 12, 0, 0),
    )
    assert report.rows_added == 2
    memo = report.memos[0]
    memo_ids = _index_rows(report.workbook_path)
    assert memo.memo_id in memo_ids
    assert f"{memo.memo_id}-A2" in memo_ids

    workbook = openpyxl.load_workbook(report.workbook_path, read_only=True)
    sheet = workbook["Index"]
    by_memo_id: dict[str, tuple] = {}
    for row in sheet.iter_rows(min_row=4, max_col=48, values_only=True):
        if row[0]:
            by_memo_id[str(row[0])] = row
    workbook.close()
    first_row = by_memo_id[memo.memo_id]
    second_row = by_memo_id[f"{memo.memo_id}-A2"]
    assert first_row[10] == "Broadband Partners"  # col 11: Portfolio Company
    assert second_row[10] == "GridCo Transmission"
    assert first_row[47] == 120.0  # col 48: Fund Share Equity Value ($M)
    assert second_row[47] == 95.0
    # booleans land as Yes/No strings, never formulas
    assert first_row[43] in (None, "Yes", "No")


# --------------------------------------------------------------------------
# Review Flags sheet schema + Run Log row
# --------------------------------------------------------------------------


def test_review_flags_row_matches_13_column_schema(phase2_env) -> None:
    report = run(
        phase2_env, scope="deal", client="Apollo Global Management", deal="Hyperoptic",
        period="Q1 2026", force=True, now=datetime(2026, 6, 12, 13, 0, 0),
    )
    memo = report.memos[0]
    rows = [r for r in _flag_rows(report.workbook_path) if r[0] == report.run_id]
    assert rows, "no flag rows written for this run"
    row = rows[0]
    assert len(row) == 13
    run_id, memo_id, filename, manager, company, val_date, qa, flag_no, desc, cat, attn, resolved, notes = row
    assert run_id == report.run_id and memo_id == memo.memo_id
    assert filename == memo.file_name
    assert manager == "Apollo Global Management" and company == "Hyperoptic"
    assert val_date == "2026-03-31" and qa == "qa_pass_with_flags"
    assert flag_no == 1 and "image-based table" in desc and cat == "reader"
    assert attn == "Y" and resolved == "N" and notes is None


def test_run_log_row_appended(phase2_env) -> None:
    report = run(
        phase2_env, scope="deal", client="Angelo Gordon", deal="T.D. Williamson",
        period="2025-12-31", force=True, now=datetime(2026, 6, 12, 14, 0, 0),
    )
    workbook = openpyxl.load_workbook(report.workbook_path, read_only=True)
    sheet = workbook["Run Log"]
    header = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
    last = list(sheet.iter_rows(min_row=2, values_only=True))[-1]
    workbook.close()
    assert header == RUN_LOG_COLUMNS
    record = dict(zip(header, last))
    assert record["Run ID"] == report.run_id
    assert record["Run Date"] == "2026-06-12"
    assert record["Memos Processed"] == 1
    assert record["Assets Extracted"] == 1
    assert record["Records Added to Index"] == 1
    assert record["Batch Sessions"] is None  # Phase 3
    assert "LLM fallback disabled" in record["Notes"]


def test_reference_template_never_modified(project_root: Path, phase2_env) -> None:
    """The run above copied the reference workbook; the original must be
    byte-identical to its committed state (rule: NEVER touch the template)."""
    import hashlib

    digest_before = hashlib.sha256(REFERENCE_TEMPLATE.read_bytes()).hexdigest()
    run(
        phase2_env, scope="deal", client="Angelo Gordon", deal="Accell",
        period="2025-01-31", force=True, now=datetime(2026, 6, 12, 15, 0, 0),
    )
    assert hashlib.sha256(REFERENCE_TEMPLATE.read_bytes()).hexdigest() == digest_before


# --------------------------------------------------------------------------
# Phase 4: review-queue writer entry points (single-cell update, flag resolve)
# --------------------------------------------------------------------------


def test_update_cell_overwrites_one_cell_by_col_index(phase2_env) -> None:
    report = run(
        phase2_env, scope="deal", client="Angelo Gordon", deal="T.D. Williamson",
        period="2025-12-31", force=True, now=datetime(2026, 6, 12, 16, 0, 0),
    )
    memo = report.memos[0]
    row_memo_id = memo.assets[0].row_memo_id
    schema = _schema_fields(Path(__file__).resolve().parent.parent)
    target = next(f for f in schema if f.header == "Portfolio Company")

    writer = WorkbookWriter(report.workbook_path, schema, pv_root=phase2_env.pv_root)
    before_rows = _index_rows(report.workbook_path)
    row = writer.update_cell(row_memo_id, target.col_index, "Corrected Co")
    writer.save()

    workbook = openpyxl.load_workbook(report.workbook_path, read_only=True)
    sheet = workbook["Index"]
    assert sheet.cell(row=row, column=target.col_index).value == "Corrected Co"
    assert sheet.cell(row=row, column=1).value == row_memo_id  # row identity intact
    workbook.close()
    assert _index_rows(report.workbook_path) == before_rows  # no rows added/removed


def test_update_cell_value_conventions_and_formula_defused(phase2_env) -> None:
    report = run(
        phase2_env, scope="deal", client="Angelo Gordon", deal="T.D. Williamson",
        period="2025-12-31", force=True, now=datetime(2026, 6, 12, 16, 30, 0),
    )
    row_memo_id = report.memos[0].assets[0].row_memo_id
    schema = _schema_fields(Path(__file__).resolve().parent.parent)
    target = next(f for f in schema if f.header == "Portfolio Company")

    writer = WorkbookWriter(report.workbook_path, schema, pv_root=phase2_env.pv_root)
    row = writer.update_cell(row_memo_id, target.col_index, True)
    writer.update_cell(row_memo_id, target.col_index + 1, "=SUM(A1:A9)")
    writer.save()

    workbook = openpyxl.load_workbook(report.workbook_path)
    sheet = workbook["Index"]
    assert sheet.cell(row=row, column=target.col_index).value == "Yes"
    cell = sheet.cell(row=row, column=target.col_index + 1)
    assert cell.value == "=SUM(A1:A9)" and cell.data_type == "s"  # string, never formula
    workbook.close()


def test_update_cell_unknown_memo_raises(phase2_env, tmp_path: Path, project_root: Path) -> None:
    copy = tmp_path / "copy.xlsx"
    copy_template(REFERENCE_TEMPLATE, copy, pv_root=str(tmp_path / "fake_pv"))
    writer = WorkbookWriter(copy, _schema_fields(project_root), pv_root=str(tmp_path / "fake_pv"))
    with pytest.raises(KeyError, match="MEMO_NOPE"):
        writer.update_cell("MEMO_NOPE", 5, "x")


def test_resolve_flag_marks_row_and_sets_note(phase2_env) -> None:
    report = run(
        phase2_env, scope="deal", client="Apollo Global Management", deal="Hyperoptic",
        period="Q1 2026", force=True, now=datetime(2026, 6, 12, 17, 0, 0),
    )
    flagged = next(
        (asset, flag)
        for memo in report.memos
        for asset in memo.assets
        for flag in asset.flags
    )
    asset, flag = flagged
    schema = _schema_fields(Path(__file__).resolve().parent.parent)
    writer = WorkbookWriter(report.workbook_path, schema, pv_root=phase2_env.pv_root)
    assert writer.resolve_flag(asset.row_memo_id, flag.description, resolved=True, note="checked")
    assert not writer.resolve_flag(asset.row_memo_id, "no such flag text", resolved=True)
    writer.save()

    rows = _flag_rows(report.workbook_path)
    match = next(r for r in rows if r[1] == asset.row_memo_id and r[8] == flag.description)
    assert match[11] == "Y" and match[12] == "checked"
